"""Tests for the calibration-v3 owner-vs-GC split on `BidPackage`.

Calibration v3 surfaced that `BidPackage.contractor` was conflating two
distinct entities: the *owning agency / institution* paying for the work
(USFWS, ASU, TAMU) and the *general contractor / construction manager*
running the bid (Beck, JE Dunn, SSC). On government direct solicitations
there is no intermediate GC — the owner is the agency and `gc` is None.

These tests confirm:
  * the new `owner` and `gc` fields exist with sane defaults
  * a model_validator mirrors `contractor` <-> `gc` so legacy payloads
    (LLM emits only `contractor`) and new payloads (LLM emits only `gc`)
    both round-trip cleanly without losing data
  * `_coerce_bid_package` reads all three fields out of the LLM response
    and respects the legacy fallback
  * the Excel and PDF exports survive a BidPackage with owner+gc set
    without crashing (round-trip smoke)

No real LLM calls — the extractor is exercised with a hand-rolled dict.
"""

from __future__ import annotations

from pathlib import Path

from core.extractors import _coerce_bid_package
from core.exporter import export_estimate_xlsx
from core.exporter_pdf import build_quote_pdf
from core.schemas import (
    BidPackage,
    Estimate,
    PaymentMilestone,
    PaymentSchedule,
    QuoteConfig,
    QuoteMeta,
    SiteInfo,
)
from core.takeoff import ProjectInfo, ProjectModel, ScopeMatrix


# ---------------------------------------------------------------------------
# Validator behavior
# ---------------------------------------------------------------------------


def test_legacy_contractor_copied_into_gc() -> None:
    """Pre-v3 payload: `contractor` set, `gc` None — validator backfills `gc`."""
    bp = BidPackage(
        pdf_name="usfws_san_marcos_solicitation.pdf",
        trade_name=None,
        contractor="U.S. Fish & Wildlife Service",
    )
    assert bp.contractor == "U.S. Fish & Wildlife Service"
    assert bp.gc == "U.S. Fish & Wildlife Service", \
        "validator must mirror contractor -> gc on legacy payloads"
    assert bp.owner is None, \
        "owner stays None on legacy payloads — LLM didn't emit it"
    assert bp.document_kind == "trade_package", "default document_kind"


def test_new_gc_only_mirrors_back_into_contractor() -> None:
    """New-style payload: `gc` set, `contractor` None — validator backfills."""
    bp = BidPackage(
        pdf_name="beck_drywall_package.pdf",
        trade_name="Drywall",
        owner="Dallas ISD",
        gc="BECK 3I JOINT VENTURE",
    )
    assert bp.gc == "BECK 3I JOINT VENTURE"
    assert bp.contractor == "BECK 3I JOINT VENTURE", \
        "validator must mirror gc -> contractor for backcompat"
    assert bp.owner == "Dallas ISD"


def test_owner_only_government_direct_solicitation() -> None:
    """Direct gov-to-sub: only `owner` set, `gc` and `contractor` stay None."""
    bp = BidPackage(
        pdf_name="esbd_516718_carr_efa.pdf",
        trade_name=None,
        owner="Angelo State University",
        document_kind="trade_package",
    )
    assert bp.owner == "Angelo State University"
    assert bp.gc is None
    assert bp.contractor is None


def test_both_gc_and_contractor_set_no_overwrite() -> None:
    """When both are set explicitly (mixed-source payload), validator is a no-op."""
    bp = BidPackage(
        pdf_name="x.pdf",
        gc="JE Dunn",
        contractor="Old Stale Name",
    )
    # Validator never clobbers user-supplied values.
    assert bp.gc == "JE Dunn"
    assert bp.contractor == "Old Stale Name"


# ---------------------------------------------------------------------------
# _coerce_bid_package — pulls owner/gc/contractor from the LLM response
# ---------------------------------------------------------------------------


def test_coerce_reads_both_owner_and_gc() -> None:
    """LLM emits both owner and gc — both populated, contractor mirrored."""
    data = {
        "trade_name": "Concrete",
        "owner": "Dallas ISD",
        "gc": "BECK 3I JOINT VENTURE",
    }
    bp = _coerce_bid_package(data, "beck_concrete.pdf")
    assert bp is not None
    assert bp.owner == "Dallas ISD"
    assert bp.gc == "BECK 3I JOINT VENTURE"
    assert bp.contractor == "BECK 3I JOINT VENTURE", \
        "validator copies gc -> contractor for backward compat"


def test_coerce_legacy_only_contractor() -> None:
    """LLM emits only legacy `contractor` — gc backfilled, owner None."""
    data = {
        "trade_name": None,
        "contractor": "Angelo State University",
    }
    bp = _coerce_bid_package(data, "asu_legacy.pdf")
    assert bp is not None
    assert bp.contractor == "Angelo State University"
    assert bp.gc == "Angelo State University", \
        "coerce + validator together must promote legacy contractor to gc"
    assert bp.owner is None


def test_coerce_owner_only_no_gc() -> None:
    """LLM emits only `owner` (direct gov solicitation)."""
    data = {
        "trade_name": None,
        "owner": "U.S. Fish & Wildlife Service",
    }
    bp = _coerce_bid_package(data, "usfws_direct.pdf")
    assert bp is not None
    assert bp.owner == "U.S. Fish & Wildlife Service"
    assert bp.gc is None
    assert bp.contractor is None


def test_coerce_document_kind_defaults_to_trade_package() -> None:
    """Missing or invalid `document_kind` defaults to 'trade_package'."""
    bp1 = _coerce_bid_package({"trade_name": "MEP"}, "x.pdf")
    bp2 = _coerce_bid_package({"trade_name": "MEP", "document_kind": "garbage"}, "y.pdf")
    bp3 = _coerce_bid_package({"trade_name": "MEP", "document_kind": "supporting_document"}, "z.pdf")
    assert bp1.document_kind == "trade_package"
    assert bp2.document_kind == "trade_package", "invalid value coerces to default"
    assert bp3.document_kind == "supporting_document"


# ---------------------------------------------------------------------------
# Excel + PDF round-trip — preserves both fields, no crashes
# ---------------------------------------------------------------------------


def _project_with_owner_gc() -> ProjectModel:
    pkgs = [
        BidPackage(
            pdf_name="usfws_solicitation.pdf",
            package_number="01",
            trade_name="Mechanical",
            owner="U.S. Fish & Wildlife Service",
            gc=None,
            document_kind="trade_package",
        ),
        BidPackage(
            pdf_name="beck_concrete.pdf",
            package_number="03",
            trade_name="Concrete",
            owner="Dallas ISD",
            gc="BECK 3I JOINT VENTURE",
            document_kind="trade_package",
        ),
    ]
    return ProjectModel(
        rooms=[], doors=[], windows=[], structural=[], mep=[], spec_sections=[],
        site=SiteInfo(),
        takeoffs=[], sheet_summaries={}, warnings=[],
        bid_packages=pkgs,
        project_info=ProjectInfo(name="Mixed Project", number="X-1"),
        scope_matrix=ScopeMatrix(packages=pkgs, by_division={}, all_alternates=[], coverage_warnings=[]),
        aggregated_inclusions=[], aggregated_exclusions=[],
    )


def test_excel_export_includes_owner_and_gc_columns() -> None:
    """Owner + General Contractor columns survive the Excel round-trip."""
    from openpyxl import load_workbook
    from io import BytesIO

    project = _project_with_owner_gc()
    estimate = Estimate(project_name="Mixed Project")
    payload = export_estimate_xlsx(
        estimate=estimate, project=project, csi_titles={}, extractions=[],
    )
    wb = load_workbook(BytesIO(payload))
    assert "Bid Packages" in wb.sheetnames
    ws = wb["Bid Packages"]
    headers = [c.value for c in ws[1]]
    assert "Owner" in headers, f"Owner column missing from Bid Packages: {headers}"
    assert "General Contractor" in headers, \
        f"General Contractor column missing: {headers}"

    # Read back the two data rows by column name.
    owner_col = headers.index("Owner") + 1
    gc_col = headers.index("General Contractor") + 1
    row2 = [ws.cell(row=2, column=i).value for i in range(1, len(headers) + 1)]
    row3 = [ws.cell(row=3, column=i).value for i in range(1, len(headers) + 1)]
    owners = {row2[owner_col - 1], row3[owner_col - 1]}
    gcs = {row2[gc_col - 1], row3[gc_col - 1]}
    assert "U.S. Fish & Wildlife Service" in owners
    assert "Dallas ISD" in owners
    assert "BECK 3I JOINT VENTURE" in gcs
    # USFWS direct solicitation has no GC.
    assert "" in gcs or None in gcs


def test_pdf_export_with_owner_gc_smoke(tmp_path: Path) -> None:
    """PDF builder doesn't crash when packages carry owner+gc metadata.

    Also exercises a non-empty estimate so the three-tile path (not the
    new empty-state banner) gets covered here. Smoke-only — visual layout
    is reviewed in the Streamlit Client Quote tab.
    """
    from core.schemas import Alternate, CostCategory, CostLine

    project = _project_with_owner_gc()
    # Give one package an alternate so the "Alternates and unit prices"
    # section renders and exercises the owner/gc caption rendering.
    project.bid_packages[1].alternates.append(
        Alternate(number="Alt #1", description="Decorative concrete option", add_or_deduct="Add")
    )
    estimate = Estimate(
        project_name="Mixed Project",
        line_items=[
            CostLine(
                csi_division="03", csi_section="03 30 00",
                description="Slab on grade", quantity=100.0, unit="SF",
                unit_cost=10.0, total_cost=1000.0,
                cost_category=CostCategory.MATERIAL,
            ),
        ],
    )
    out = tmp_path / "owner_gc_smoke.pdf"
    result = build_quote_pdf(
        estimate=estimate,
        project=project,
        quote_config=QuoteConfig(
            quote_meta=QuoteMeta(scope_blurb="smoke"),
            payment_schedule=PaymentSchedule(
                mode="percentage",
                milestones=[PaymentMilestone(label="Done", percentage=100.0)],
            ),
        ),
        out_path=out,
    )
    assert result == out
    assert out.is_file() and out.stat().st_size > 4000
    assert out.read_bytes()[:5] == b"%PDF-"
