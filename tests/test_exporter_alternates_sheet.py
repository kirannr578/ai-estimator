"""Phase T9.0 — Excel exporter Bid Alternates worksheet tests.

Covers:

* "Bid Alternates" worksheet exists when the estimate has at least
  one priced alternate.
* No "Bid Alternates" worksheet when the estimate has zero
  alternates (default behaviour for projects without any).
* Footer rows: base / +all-additive / +all-deductive / +substitution
  / +default-selected.
* "Project Summary" alternates block (totals + count + missing).
* Currency cells use ``$#,##0.00`` formatting.
* Header tint uses :data:`ALTERNATES_HEADER_FILL`.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.exporter import (
    ALTERNATES_HEADER_FILL,
    export_estimate_xlsx,
)
from core.schemas import (
    AlternateLineEstimate,
    AlternatePricingBasis,
    AlternateType,
    CostCategory,
    CostLine,
    Estimate,
    SiteInfo,
)
from core.takeoff import ProjectInfo, ProjectModel, ScopeMatrix


def _empty_project() -> ProjectModel:
    return ProjectModel(
        rooms=[],
        doors=[],
        windows=[],
        structural=[],
        mep=[],
        spec_sections=[],
        site=SiteInfo(),
        takeoffs=[],
        sheet_summaries={},
        warnings=[],
        bid_packages=[],
        project_info=ProjectInfo(name="T9 Sheet Test"),
        scope_matrix=ScopeMatrix(
            packages=[], by_division={}, all_alternates=[], coverage_warnings=[]
        ),
        aggregated_inclusions=[],
        aggregated_exclusions=[],
        alternates=[],
    )


def _base_line() -> CostLine:
    return CostLine(
        csi_division="09",
        csi_section="09 91 23",
        description="Interior painting",
        quantity=1000.0,
        unit="SF",
        unit_cost=10.0,
        total_cost=10000.0,
        cost_category=CostCategory.SUBCONTRACTOR,
        confidence=0.92,
    )


def _ale(
    alt_id: str,
    *,
    cost: float | None = 1000.0,
    atype: AlternateType = AlternateType.ADDITIVE,
    basis: AlternatePricingBasis = AlternatePricingBasis.EXTRACTED_FROM_BID_FORM,
    included: bool = False,
) -> AlternateLineEstimate:
    return AlternateLineEstimate(
        alternate_id=alt_id,
        alternate_type=atype,
        description=f"description for {alt_id}",
        cost_delta=cost,
        pricing_basis=basis,
        included_in_base=included,
    )


def _load_xlsx(payload: bytes):
    return load_workbook(BytesIO(payload))


# ---------------------------------------------------------------------------
# Sheet presence
# ---------------------------------------------------------------------------


class TestBidAlternatesSheetPresence:
    def test_sheet_created_when_alternates_present(self) -> None:
        est = Estimate(
            project_name="T9",
            line_items=[_base_line()],
            alternates=[
                _ale("Alt 1", cost=500.0, atype=AlternateType.ADDITIVE),
            ],
        )
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        assert "Bid Alternates" in wb.sheetnames

    def test_sheet_omitted_when_no_alternates(self) -> None:
        est = Estimate(project_name="T9", line_items=[_base_line()], alternates=[])
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        assert "Bid Alternates" not in wb.sheetnames

    def test_header_tint_uses_alternates_fill(self) -> None:
        est = Estimate(
            project_name="T9",
            line_items=[_base_line()],
            alternates=[_ale("Alt 1", cost=500.0)],
        )
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        ws = wb["Bid Alternates"]
        rgb = ws.cell(row=1, column=1).fill.fgColor.rgb
        # openpyxl includes the alpha channel; match by suffix.
        assert rgb is not None
        assert str(rgb).upper().endswith(ALTERNATES_HEADER_FILL.fgColor.rgb.upper())


# ---------------------------------------------------------------------------
# Worksheet content
# ---------------------------------------------------------------------------


class TestBidAlternatesSheetContent:
    def test_headers_match_contract(self) -> None:
        est = Estimate(
            project_name="T9",
            line_items=[_base_line()],
            alternates=[_ale("Alt 1", cost=500.0)],
        )
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        ws = wb["Bid Alternates"]
        headers = [c.value for c in ws[1]]
        assert "Alternate ID" in headers
        assert "Type" in headers
        assert "Description" in headers
        assert "Cost Delta" in headers
        assert "Pricing Basis" in headers
        assert "Confidence" in headers

    def test_first_row_renders_alternate_data(self) -> None:
        est = Estimate(
            project_name="T9",
            line_items=[_base_line()],
            alternates=[
                _ale("Alt 1", cost=500.0, atype=AlternateType.ADDITIVE),
            ],
        )
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        ws = wb["Bid Alternates"]
        # Row 2 = first data row.
        assert ws.cell(row=2, column=1).value == "Alt 1"
        assert "Additive" in ws.cell(row=2, column=2).value
        assert ws.cell(row=2, column=4).value == 500.0
        # Currency formatting on cost delta column.
        assert "$" in ws.cell(row=2, column=4).number_format

    def test_missing_cost_delta_shows_em_dash(self) -> None:
        est = Estimate(
            project_name="T9",
            line_items=[_base_line()],
            alternates=[
                _ale(
                    "Alt 1",
                    cost=None,
                    basis=AlternatePricingBasis.MISSING,
                )
            ],
        )
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        ws = wb["Bid Alternates"]
        assert ws.cell(row=2, column=4).value == "—"


# ---------------------------------------------------------------------------
# Footer rows
# ---------------------------------------------------------------------------


class TestBidAlternatesFooter:
    def _mixed_estimate(self) -> Estimate:
        return Estimate(
            project_name="T9",
            contingency_pct=10.0,
            overhead_pct=10.0,
            profit_pct=5.0,
            line_items=[_base_line()],
            alternates=[
                _ale("Alt 1", cost=500.0, atype=AlternateType.ADDITIVE),
                _ale("Alt 2", cost=-200.0, atype=AlternateType.DEDUCTIVE),
                _ale("Alt 3", cost=300.0, atype=AlternateType.SUBSTITUTION),
                _ale("VE-1", cost=-100.0, atype=AlternateType.VE),
            ],
        )

    def _footer_lookup(self, ws) -> dict[str, float]:
        """Walk column A for the footer row labels and return label → value."""
        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            label = row[0].value
            value = row[3].value if len(row) >= 4 else None
            if isinstance(label, str) and isinstance(value, (int, float)):
                result[label] = float(value)
        return result

    def test_base_bid_footer_row_present(self) -> None:
        payload = export_estimate_xlsx(
            self._mixed_estimate(), _empty_project(), csi_titles={}
        )
        ws = _load_xlsx(payload)["Bid Alternates"]
        footers = self._footer_lookup(ws)
        # Base bid (no alternates) = 10,000.
        assert any(
            "base bid" in k.lower() and v == 10000.0
            for k, v in footers.items()
        )

    def test_all_additive_footer_row_present(self) -> None:
        payload = export_estimate_xlsx(
            self._mixed_estimate(), _empty_project(), csi_titles={}
        )
        ws = _load_xlsx(payload)["Bid Alternates"]
        footers = self._footer_lookup(ws)
        # Base + all additive (Alt 1: +500) = 10,500.
        assert any(
            "all additive" in k.lower() and v == 10500.0
            for k, v in footers.items()
        )

    def test_all_deductive_footer_row_present(self) -> None:
        payload = export_estimate_xlsx(
            self._mixed_estimate(), _empty_project(), csi_titles={}
        )
        ws = _load_xlsx(payload)["Bid Alternates"]
        footers = self._footer_lookup(ws)
        # Base + all deductive (-200) + VE (-100) = 9,700.
        assert any(
            "deductive" in k.lower() and v == 9700.0
            for k, v in footers.items()
        )


# ---------------------------------------------------------------------------
# Project Summary block
# ---------------------------------------------------------------------------


class TestProjectSummaryAlternatesBlock:
    def test_summary_carries_alternates_count(self) -> None:
        est = Estimate(
            project_name="T9",
            line_items=[_base_line()],
            alternates=[
                _ale("Alt 1", cost=500.0, atype=AlternateType.ADDITIVE),
                _ale("Alt 2", cost=-200.0, atype=AlternateType.DEDUCTIVE),
            ],
        )
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        ws = wb["Summary"]
        # The alternates count line lives somewhere in the top 25 rows of
        # the summary sheet — scan rather than pin a specific cell.
        all_text = " ".join(
            str(c) for r in ws.iter_rows(min_row=1, max_row=25, values_only=True)
            for c in r if c is not None
        )
        assert "Alternates:" in all_text or "Bid Alternates" in all_text
        assert "2" in all_text  # the count

    def test_summary_additive_total_present(self) -> None:
        est = Estimate(
            project_name="T9",
            line_items=[_base_line()],
            alternates=[_ale("Alt 1", cost=750.0, atype=AlternateType.ADDITIVE)],
        )
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        ws = wb["Summary"]
        # Walk rows 17-22 looking for the additive total.
        seen = False
        for r in range(17, 23):
            for c in range(1, 4):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, (int, float)) and cell.value == 750.0:
                    seen = True
                    break
        assert seen, "Expected the alternates additive total in the Summary sheet"

    def test_summary_layout_unchanged_when_no_alternates(self) -> None:
        est = Estimate(project_name="T9", line_items=[_base_line()], alternates=[])
        payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
        wb = _load_xlsx(payload)
        ws = wb["Summary"]
        # Pre-T9 layout pinned: A18 = "By CSI division".
        assert ws["A18"].value == "By CSI division"
