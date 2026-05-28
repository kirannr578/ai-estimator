"""Phase T6.4.c — Excel exporter "Notes" column visibility tests.

Confirms the Excel exporter's "Line Items" (a.k.a. Cost Estimate) sheet
surfaces every :pyattr:`~core.schemas.CostLine.notes` value verbatim
into the Notes column, and that the column is wide enough that a
typical ``[sub-quote] operator override: ...`` provenance tag remains
visible at a glance when the workbook is opened by an estimator
without manual column re-sizing.

The five tests below ride on top of the T6.4.c source-tag-propagation
contract pinned in :mod:`tests.test_subquote_tag_propagation`; this
file is the exporter-side companion that confirms the propagated tag
actually surfaces in the deliverable.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.exporter import export_estimate_xlsx
from core.pricing.batch_override import (
    SOURCE_TAG_PATTERN,
    SOURCE_TAG_SUBQUOTE_LLM,
    SOURCE_TAG_SUBQUOTE_TABULAR,
    SOURCE_TAG_VENDOR_CSV,
    BatchMatchResult,
    BatchMatchStatus,
    BatchOverridePlan,
    BatchOverrideRow,
    apply_batch_plan,
)
from core.schemas import (
    CostBand,
    CostCategory,
    CostLine,
    CostSourceTier,
    Estimate,
    SiteInfo,
)
from core.takeoff import ProjectInfo, ProjectModel, ScopeMatrix


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _empty_project() -> ProjectModel:
    return ProjectModel(
        rooms=[],
        doors=[],
        windows=[],
        structural=[],
        mep=[],
        spec_sections=[],
        site=SiteInfo(),
        takeoffs=[],
        sheet_summaries={},
        warnings=[],
        bid_packages=[],
        project_info=ProjectInfo(name="T6.4.c Notes Visibility"),
        scope_matrix=ScopeMatrix(
            packages=[], by_division={}, all_alternates=[], coverage_warnings=[]
        ),
        aggregated_inclusions=[],
        aggregated_exclusions=[],
    )


def _line(
    *,
    description: str = "Interior latex paint walls",
    csi_division: str = "09",
    csi_section: str = "09 91 23",
    quantity: float = 100.0,
    unit: str = "SF",
    unit_cost: float = 2.0,
    notes: str | None = None,
) -> CostLine:
    return CostLine(
        csi_division=csi_division,
        csi_section=csi_section,
        description=description,
        quantity=quantity,
        unit=unit,
        unit_cost=unit_cost,
        total_cost=round(unit_cost * quantity, 2),
        cost_category=CostCategory.SUBCONTRACTOR,
        confidence=0.92,
        price_confidence=0.65,
        cost_source_tier=CostSourceTier.INTERPOLATED,
        cost_band=CostBand.OPERATOR_REVIEW,
        suppressed=False,
        cost_source="cwicr:42",
        notes=notes,
    )


def _estimate(lines: list[CostLine]) -> Estimate:
    return Estimate(
        project_name="T6.4.c Notes Visibility",
        region_multiplier=1.0,
        contingency_pct=10.0,
        overhead_pct=10.0,
        profit_pct=5.0,
        line_items=lines,
    )


def _apply_subquote_to_estimate(
    estimate: Estimate, *, source_tag: str
) -> Estimate:
    """Apply a single-row MATCHED plan to the first line of ``estimate``."""
    row = BatchOverrideRow(
        row_index=2,
        description=estimate.line_items[0].description,
        unit_cost=3.50,
        vendor="Sherwin",
        quote_ref="Q-1",
    )
    plan = BatchOverridePlan(
        total_rows=1,
        matched=[
            BatchMatchResult(
                row=row,
                status=BatchMatchStatus.MATCHED,
                best_match_index=0,
                best_match_similarity=1.0,
                runner_up_index=None,
                runner_up_similarity=0.0,
                candidate_lines=[(0, 1.0)],
            )
        ],
        ambiguous=[],
        no_match=[],
        low_similarity=[],
        similarity_threshold=0.65,
        ambiguity_margin=0.10,
    )
    new_estimate, _ = apply_batch_plan(estimate, plan, source_tag=source_tag)
    return new_estimate


def _line_items_sheet(workbook):
    """Return the "Line Items" worksheet (the priced-rows sheet)."""
    assert "Line Items" in workbook.sheetnames
    return workbook["Line Items"]


def _data_row_for(sheet, description: str) -> int:
    """1-indexed row whose Description column (col 4) matches ``description``."""
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=4).value == description:
            return r
    raise AssertionError(f"row for description {description!r} not found")


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_line_items_sheet_has_notes_header_in_last_column() -> None:
    """The "Line Items" sheet's last header is "Notes".

    Pins the column-layout contract: ``Notes`` is the rightmost column
    of the Line Items / Cost Estimate sheet (column 21 in the current
    schema). A future schema change that drops the column will fail
    this test loudly.
    """
    estimate = _estimate([_line()])
    xlsx = export_estimate_xlsx(
        estimate, _empty_project(), csi_titles={"09": "Finishes"}
    )
    wb = load_workbook(BytesIO(xlsx), read_only=False, data_only=False)
    ws = _line_items_sheet(wb)
    # Header row (row 1) — find the rightmost non-empty header.
    headers = [ws.cell(row=1, column=c).value for c in range(1, 25) if ws.cell(row=1, column=c).value]
    assert headers[-1] == "Notes"


def test_notes_cell_value_equals_costline_notes_field() -> None:
    """``CostLine.notes`` round-trips into the Notes cell verbatim.

    No truncation, no reformatting — whatever the source-tag rewrite
    helper produced is exactly what the exporter writes into the
    workbook cell.
    """
    estimate = _apply_subquote_to_estimate(
        _estimate([_line()]), source_tag=SOURCE_TAG_SUBQUOTE_TABULAR
    )
    expected = estimate.line_items[0].notes
    assert expected is not None
    xlsx = export_estimate_xlsx(
        estimate, _empty_project(), csi_titles={"09": "Finishes"}
    )
    wb = load_workbook(BytesIO(xlsx), read_only=False, data_only=False)
    ws = _line_items_sheet(wb)
    r = _data_row_for(ws, "Interior latex paint walls")
    notes_cell = ws.cell(row=r, column=21).value
    assert notes_cell == expected


def test_notes_cell_starts_with_source_tag_for_each_source() -> None:
    """For every batch-applied source, the Notes cell starts with the tag.

    Confirms the T6.4.c contract end-to-end: ``apply_batch_plan`` →
    rewritten ``CostLine.notes`` → Excel exporter Notes column starts
    with ``[<source-tag>] ``. Three sources tested: ``[vendor-csv]``,
    ``[sub-quote]``, ``[sub-quote-llm]``.
    """
    import re

    pat = re.compile(SOURCE_TAG_PATTERN)
    for tag in (
        SOURCE_TAG_VENDOR_CSV,
        SOURCE_TAG_SUBQUOTE_TABULAR,
        SOURCE_TAG_SUBQUOTE_LLM,
    ):
        estimate = _apply_subquote_to_estimate(
            _estimate([_line()]), source_tag=tag
        )
        xlsx = export_estimate_xlsx(
            estimate, _empty_project(), csi_titles={"09": "Finishes"}
        )
        wb = load_workbook(BytesIO(xlsx), read_only=False, data_only=False)
        ws = _line_items_sheet(wb)
        r = _data_row_for(ws, "Interior latex paint walls")
        notes_cell = ws.cell(row=r, column=21).value
        assert notes_cell is not None
        assert notes_cell.startswith(tag + " "), (
            f"tag {tag} not at start of {notes_cell!r}"
        )
        assert pat.match(notes_cell) is not None


def test_notes_column_width_adequate_for_tag_plus_summary() -> None:
    """Notes column auto-sizes wide enough to show the source tag.

    The ``_autosize`` helper widens each column to ``min(max(longest+2, 10),
    max_width)``. Even at ``max_width=60``, a ``[sub-quote-llm] operator
    override: ...`` cell easily exceeds 30 chars and clamps the width
    against the upper bound. The pin: the Notes column width must be
    at least ``len("[sub-quote-llm] operator override: ") = 35`` so the
    provenance is visible without manual resizing.
    """
    estimate = _apply_subquote_to_estimate(
        _estimate([_line()]), source_tag=SOURCE_TAG_SUBQUOTE_LLM
    )
    xlsx = export_estimate_xlsx(
        estimate, _empty_project(), csi_titles={"09": "Finishes"}
    )
    wb = load_workbook(BytesIO(xlsx), read_only=False, data_only=False)
    ws = _line_items_sheet(wb)
    # Column 21 = Notes; openpyxl exposes the letter via the COLUMN_LETTER
    # mapping. For column 21 the letter is "U".
    notes_col_letter = "U"
    width = ws.column_dimensions[notes_col_letter].width
    assert width is not None, "Notes column has no width set"
    # Minimum bar: enough chars to show "[sub-quote-llm] operator override:"
    # (35 chars) so the tag + sentinel are visible at a glance.
    assert width >= 35, (
        f"Notes column width {width} too narrow for tag-plus-sentinel"
    )


def test_multi_line_estimate_every_notes_row_shows_tag_at_start() -> None:
    """Multi-line estimate: every applied line's Notes cell starts with a tag.

    Pins the end-to-end T6.4.c contract across multiple CostLines: the
    rewrite helper does NOT regress on the first line because of the
    second, and vice versa. Two-line estimate; both applied via
    ``apply_batch_plan`` with the same source tag.
    """
    import re

    estimate = _estimate([
        _line(description="Lavatory P-1", unit_cost=500.0),
        _line(description="Water closet WC-1", unit_cost=400.0),
    ])
    rows = [
        BatchOverrideRow(2, "Lavatory P-1", 450.0),
        BatchOverrideRow(3, "Water closet WC-1", 525.0),
    ]
    from core.pricing.batch_override import match_cost_lines
    plan = match_cost_lines(rows, list(estimate.line_items))
    new_estimate, _ = apply_batch_plan(
        estimate, plan, source_tag=SOURCE_TAG_SUBQUOTE_LLM
    )
    xlsx = export_estimate_xlsx(
        new_estimate, _empty_project(), csi_titles={"09": "Finishes"}
    )
    wb = load_workbook(BytesIO(xlsx), read_only=False, data_only=False)
    ws = _line_items_sheet(wb)
    pat = re.compile(SOURCE_TAG_PATTERN)

    for desc in ("Lavatory P-1", "Water closet WC-1"):
        r = _data_row_for(ws, desc)
        notes_cell = ws.cell(row=r, column=21).value
        assert notes_cell is not None, f"empty Notes cell for {desc}"
        assert notes_cell.startswith(SOURCE_TAG_SUBQUOTE_LLM + " "), (
            f"tag missing from {desc} notes: {notes_cell!r}"
        )
        assert pat.match(notes_cell) is not None
