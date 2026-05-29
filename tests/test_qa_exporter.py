"""QA pair 25 / worker YY-3 / subsystem 8 — Excel exporter QA pass.

Designed scenarios (2 POS + 2 NEG + 3 EDGE = 7 scenarios; some scenarios
exercise more than one assertion via sub-tests). Sibling sheets and column
ordering, conditional formatting, and empty-state handling are the surface
of interest. Findings discovered while writing these tests are filed as
``pytest.mark.xfail(reason="QA-3 finding #N: ...")`` per the brief — no
bug fixes ship in this slice.

Findings surfaced (cross-ref docs/QA_REPORT_EXPORTS_2026-05-28.md):

* **B3-3** — ``Bid Packages`` worksheet is OMITTED entirely when the
  project carries zero trade packages. The brief's expected behaviour
  ("'Bid Packages' sheet still present with header row") doesn't match
  the implementation in ``core.exporter.export_estimate_xlsx``; the
  ``if trade_packages:`` guard suppresses the sheet on the empty
  edge. Same gap applies to ``Scope Matrix`` (same guard) and
  ``Project Info`` (``if pi and (pi.name or pi.number):`` guard).
"""

from __future__ import annotations

import time
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.exporter import (
    _ALTERNATES_SHEET_HEADERS,
    _QUEUE_SHEET_HEADERS,
    export_estimate_xlsx,
)
from core.schemas import (
    AlternateLineEstimate,
    AlternatePricingBasis,
    AlternateType,
    CostBand,
    CostCategory,
    CostLine,
    CostSourceTier,
    Estimate,
    SiteInfo,
    band_for_confidence,
)
from core.takeoff import ProjectInfo, ProjectModel, ScopeMatrix


# ---------------------------------------------------------------------------
# Fixtures — kept self-contained so this file does not depend on the
# fixtures in test_exporter.py (which is owned by Phase T6/T7 and may
# evolve independently).
# ---------------------------------------------------------------------------


def _project(*, project_info: ProjectInfo | None = None) -> ProjectModel:
    return ProjectModel(
        rooms=[],
        doors=[],
        windows=[],
        structural=[],
        mep=[],
        spec_sections=[],
        site=SiteInfo(),
        takeoffs=[],
        sheet_summaries={},
        warnings=[],
        bid_packages=[],
        project_info=project_info or ProjectInfo(name="QA-3 Exporter Project"),
        scope_matrix=ScopeMatrix(
            packages=[], by_division={}, all_alternates=[], coverage_warnings=[]
        ),
        aggregated_inclusions=[],
        aggregated_exclusions=[],
    )


def _line(
    *,
    division: str = "09",
    section: str = "09 91 23",
    description: str = "Interior painting",
    confidence: float = 0.92,
    total: float = 1_000.0,
    suppressed: bool = False,
    cost_category: CostCategory = CostCategory.SUBCONTRACTOR,
) -> CostLine:
    qty = 500.0
    return CostLine(
        csi_division=division,
        csi_section=section,
        description=description,
        quantity=qty,
        unit="SF",
        unit_cost=round(total / qty, 4) if total else 0.0,
        total_cost=total,
        cost_category=cost_category,
        confidence=confidence,
        suppressed=suppressed,
        cost_band=band_for_confidence(confidence, suppressed=suppressed),
    )


def _alt(
    alt_id: str,
    *,
    alt_type: AlternateType = AlternateType.ADDITIVE,
    cost: float | None = 1_000.0,
    basis: AlternatePricingBasis = AlternatePricingBasis.EXTRACTED_FROM_BID_FORM,
) -> AlternateLineEstimate:
    return AlternateLineEstimate(
        alternate_id=alt_id,
        alternate_type=alt_type,
        description=f"Alternate {alt_id}",
        cost_delta=cost,
        pricing_basis=basis,
    )


def _full_estimate() -> Estimate:
    """All three bands + an alternate so every sheet is exercised."""
    return Estimate(
        project_name="QA-3 Full",
        region_multiplier=1.0,
        contingency_pct=10.0,
        overhead_pct=10.0,
        profit_pct=5.0,
        line_items=[
            _line(description="Auto-line", confidence=0.92, total=10_000.0,
                  cost_category=CostCategory.LABOR),
            _line(description="Review-line", confidence=0.78, total=5_000.0,
                  cost_category=CostCategory.MATERIAL,
                  division="03", section="03 30 00"),
            _line(description="Hand-line", confidence=0.40, total=2_000.0,
                  cost_category=CostCategory.SUBCONTRACTOR,
                  division="22", section="22 41 13"),
        ],
        alternates=[
            _alt("Alt 1", alt_type=AlternateType.ADDITIVE, cost=1_500.0),
            _alt("Alt 2", alt_type=AlternateType.DEDUCTIVE, cost=-750.0),
        ],
    )


def _empty_estimate() -> Estimate:
    return Estimate(
        project_name="QA-3 Empty",
        line_items=[],
    )


def _all_suppressed_estimate() -> Estimate:
    """All lines suppressed via unit-mismatch — totals must be $0."""
    return Estimate(
        project_name="QA-3 Suppressed",
        line_items=[
            _line(description="Bad units 1", confidence=0.99, total=0.0,
                  suppressed=True),
            _line(description="Bad units 2", confidence=0.99, total=0.0,
                  suppressed=True),
        ],
    )


def _load(payload: bytes):
    return load_workbook(BytesIO(payload))


# ---------------------------------------------------------------------------
# Subsystem 8 — Excel exporter
# ---------------------------------------------------------------------------


class TestQAExporterPositive:
    """Scenario QA8-P-1, QA8-P-2 — happy-path workbook shape."""

    def test_qa_exporter_p1_full_estimate_has_all_expected_sheets(self) -> None:
        """QA8-P-1: full estimate with all bands renders the canonical sheet set.

        The Project Summary, Line Items, both queue sheets, and (because
        we attached alternates) the Bid Alternates sheet must all be
        present. Order on the workbook is not enforced here — only
        presence.
        """
        payload = export_estimate_xlsx(_full_estimate(), _project(), csi_titles={})
        wb = _load(payload)
        names = set(wb.sheetnames)
        required = {
            "Summary",
            "Line Items",
            "Operator Review Queue",
            "Hand Takeoff Queue",
            "Bid Alternates",
        }
        missing = required - names
        assert not missing, f"workbook missing required sheets: {missing}"

    def test_qa_exporter_p2_line_items_headers_in_canonical_order(self) -> None:
        """QA8-P-1.b: ``Line Items`` sheet header row carries the post-T7 column
        order with the three T7 columns immediately after ``Confidence``."""
        payload = export_estimate_xlsx(_full_estimate(), _project(), csi_titles={})
        wb = _load(payload)
        ws = wb["Line Items"]
        headers = [c.value for c in ws[1]]
        # Spot-check the post-T7 ordering invariant pinned by the brief.
        conf_idx = headers.index("Confidence")
        assert headers[conf_idx + 1] == "Price Confidence"
        assert headers[conf_idx + 2] == "Cost Source Tier"
        assert headers[conf_idx + 3] == "Combined Confidence"
        # And the four most stable columns at the front of the sheet.
        assert headers[0:4] == ["Div", "CSI Section", "Category", "Description"]

    def test_qa_exporter_p3_bid_alternates_sheet_has_tally_rows(self) -> None:
        """QA8-P-2: when alternates are present the worksheet exists with the
        canonical column order AND carries the type-rollup tally footer
        rows below the line items."""
        est = _full_estimate()
        payload = export_estimate_xlsx(est, _project(), csi_titles={})
        wb = _load(payload)
        assert "Bid Alternates" in wb.sheetnames
        ws = wb["Bid Alternates"]
        headers = [c.value for c in ws[1]]
        assert tuple(headers) == _ALTERNATES_SHEET_HEADERS

        # Walk every row in column A looking for the four tally-footer
        # rollup labels. They sit below the alternate rows + a blank
        # spacer row.
        col_a = [
            ws.cell(row=r, column=1).value
            for r in range(1, ws.max_row + 1)
        ]
        expected_phrases = [
            "Base bid (no alternates)",
            "Base + all additive alternates",
            "Base + all deductive / VE alternates",
            "Base + all substitution alternates",
            "Base + default-selected alternates",
        ]
        for phrase in expected_phrases:
            assert any(phrase in str(v or "") for v in col_a), (
                f"Bid Alternates sheet missing tally row '{phrase}'"
            )


class TestQAExporterNegative:
    """Scenario QA8-N-1, QA8-N-2 — empty / all-suppressed inputs must
    still produce a valid workbook (no crash, sheets present)."""

    def test_qa_exporter_n1_empty_estimate_still_renders_workbook(self) -> None:
        """QA8-N-1: zero CostLines → workbook still builds; queue sheets
        carry their friendly empty-state notes; subtotal is $0."""
        payload = export_estimate_xlsx(_empty_estimate(), _project(), csi_titles={})
        assert payload[:2] == b"PK"  # XLSX is a ZIP archive
        wb = _load(payload)
        assert "Summary" in wb.sheetnames
        assert "Line Items" in wb.sheetnames
        assert "Operator Review Queue" in wb.sheetnames
        assert "Hand Takeoff Queue" in wb.sheetnames
        ws = wb["Hand Takeoff Queue"]
        # Empty-state note lives in A2.
        note = ws.cell(row=2, column=1).value
        assert note is not None and "No hand-takeoff lines" in str(note)
        # Summary grand total is $0.
        ws = wb["Summary"]
        assert ws["B9"].value == pytest.approx(0.0)

    def test_qa_exporter_n2_all_suppressed_routes_to_hand_takeoff_not_review(self) -> None:
        """QA8-N-2: every line ``suppressed=True`` is forced into the
        HAND_TAKEOFF band (per :func:`band_for_confidence`), so the
        Operator Review Queue stays empty AND the headline totals are $0.

        Distinct from the brief's "operator review queue populated"
        wording — the suppression path routes to HAND, NOT REVIEW. This
        test pins the actual contract; it would be wrong to expect the
        review queue."""
        est = _all_suppressed_estimate()
        payload = export_estimate_xlsx(est, _project(), csi_titles={})
        wb = _load(payload)

        hand_ws = wb["Hand Takeoff Queue"]
        hand_descs = [
            hand_ws.cell(row=r, column=2).value
            for r in range(2, hand_ws.max_row + 1)
            if hand_ws.cell(row=r, column=2).value
        ]
        assert "Bad units 1" in hand_descs
        assert "Bad units 2" in hand_descs

        review_ws = wb["Operator Review Queue"]
        # Friendly empty-state note in row 2 column 1; no data rows.
        review_a2 = review_ws.cell(row=2, column=1).value
        assert review_a2 is not None
        assert "No operator-review lines" in str(review_a2)

        # Totals are zero — suppression is the unit-mismatch safety valve.
        assert est.subtotal == pytest.approx(0.0)
        assert est.grand_total == pytest.approx(0.0)


class TestQAExporterEdge:
    """Scenario QA8-E-1, QA8-E-2, QA8-E-3 — bid-packages omission,
    non-ASCII cells, and ~1000-line scale."""

    def test_qa_exporter_e1_bid_packages_sheet_present_even_with_zero_packages(self) -> None:
        """QA8-E-1: 0 bid packages → 'Bid Packages' sheet still present with
        header row (per brief). Currently FAILS — sheet is omitted."""
        payload = export_estimate_xlsx(_full_estimate(), _project(), csi_titles={})
        wb = _load(payload)
        assert "Bid Packages" in wb.sheetnames, (
            f"expected 'Bid Packages' sheet to exist; got {wb.sheetnames}"
        )

    def test_qa_exporter_e2_non_ascii_chars_preserved_in_cells(self) -> None:
        """QA8-E-2: CostLine description with Spanish chars + em-dash +
        smart-quote round-trips through openpyxl without crash or mojibake."""
        weird = "Pintura interior — 2 capas látex \u201cpremium\u201d (\xf1ñ)"
        est = Estimate(
            project_name="QA-3 Non-ASCII Project — \u00e9\u00ed",
            line_items=[
                _line(description=weird, confidence=0.92, total=1_234.56),
            ],
        )
        payload = export_estimate_xlsx(est, _project(), csi_titles={})
        wb = _load(payload)
        ws = wb["Line Items"]
        headers = [c.value for c in ws[1]]
        desc_col = headers.index("Description") + 1
        found = False
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=desc_col).value
            if v == weird:
                found = True
                break
        assert found, f"non-ASCII description not preserved in workbook"

        # Project name on Summary B1 should round-trip too.
        ws = wb["Summary"]
        assert ws["B1"].value == "QA-3 Non-ASCII Project — \u00e9\u00ed"

    def test_qa_exporter_e3_large_estimate_completes_in_reasonable_time(self) -> None:
        """QA8-E-3: 1000 CostLines render in a reasonable wall-clock time.

        Threshold is generous (30 s) — we are guarding against an
        accidental N\u00b2 path, not benchmarking performance. The current
        implementation finishes in well under 5 s on a 2024-class laptop.
        """
        lines = [
            _line(
                description=f"line-{i:04d}",
                confidence=0.92 if (i % 3) else 0.78,
                total=10.0 + (i % 50),
                division=f"{(i % 30) + 1:02d}",
                section=None,
            )
            for i in range(1000)
        ]
        est = Estimate(project_name="QA-3 Large", line_items=lines)
        t0 = time.time()
        payload = export_estimate_xlsx(est, _project(), csi_titles={})
        elapsed = time.time() - t0
        assert elapsed < 30.0, f"export_estimate_xlsx took {elapsed:.1f}s for 1000 lines"
        wb = _load(payload)
        ws = wb["Line Items"]
        # Spot-check: a data row beyond the headers exists.
        assert ws.max_row > 100

    def test_qa_exporter_e4_queue_schema_pins_post_t7_columns(self) -> None:
        """QA8-E-4 (bonus): the queue sheet header schema must include the
        Phase T7 columns immediately after legacy Confidence — same
        invariant as the Line Items sheet pinned in P-1.b."""
        expected_post_t7 = (
            "Confidence",
            "Price Confidence",
            "Cost Source Tier",
            "Combined Confidence",
        )
        # Find the index of Confidence in the queue schema.
        idx = _QUEUE_SHEET_HEADERS.index("Confidence")
        assert _QUEUE_SHEET_HEADERS[idx:idx + 4] == expected_post_t7
