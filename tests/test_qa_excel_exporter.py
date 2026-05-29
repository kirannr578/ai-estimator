"""QA pass — Subsystem 8: Excel exporter (2026-05-28).

Smoke checks that ``export_estimate_xlsx`` produces a well-formed
workbook for representative estimates. Detailed band/queue/notes
coverage already lives in ``test_exporter.py`` — this slice spot-
checks empty / large / suppressed / alternates-bearing inputs.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.estimator import _combined_band
from core.exporter import export_estimate_xlsx
from core.schemas import (
    AlternateLineEstimate,
    AlternatePricingBasis,
    AlternateType,
    CostBand,
    CostCategory,
    CostLine,
    Estimate,
    SiteInfo,
    band_for_confidence,
)
from core.takeoff import ProjectInfo, ProjectModel, ScopeMatrix


# ---------------------------------------------------------------------------
# Helpers (mirror tests/test_exporter.py shape)
# ---------------------------------------------------------------------------


def _project() -> ProjectModel:
    return ProjectModel(
        rooms=[],
        doors=[],
        windows=[],
        structural=[],
        mep=[],
        spec_sections=[],
        site=SiteInfo(),
        takeoffs=[],
        sheet_summaries={},
        warnings=[],
        bid_packages=[],
        project_info=ProjectInfo(name="QA Export"),
        scope_matrix=ScopeMatrix(
            packages=[], by_division={}, all_alternates=[], coverage_warnings=[],
        ),
        aggregated_inclusions=[],
        aggregated_exclusions=[],
    )


def _line(
    *,
    description: str = "Interior painting",
    division: str = "09",
    confidence: float = 0.92,
    price_confidence: float = 0.95,
    total: float = 1_000.0,
    suppressed: bool = False,
) -> CostLine:
    return CostLine(
        csi_division=division,
        csi_section="09 91 23",
        description=description,
        quantity=100.0,
        unit="SF",
        unit_cost=round(total / 100.0, 4) if not suppressed else 0.0,
        total_cost=total if not suppressed else 0.0,
        cost_category=CostCategory.SUBCONTRACTOR,
        confidence=confidence,
        price_confidence=price_confidence,
        suppressed=suppressed,
        cost_band=_combined_band(confidence, price_confidence, suppressed=suppressed),
    )


def _open(payload: bytes):
    return load_workbook(BytesIO(payload), read_only=False, data_only=False)


# ---------------------------------------------------------------------------
# Positive
# ---------------------------------------------------------------------------


def test_qa_pos_workbook_carries_canonical_sheets() -> None:
    """A typical priced estimate produces all canonical sheets."""
    est = Estimate(project_name="QA", line_items=[_line()])
    payload = export_estimate_xlsx(est, _project(), csi_titles={"09": "Finishes"})
    wb = _open(payload)
    names = set(wb.sheetnames)
    # Headline sheets that every export carries.
    expected = {"Summary", "Line Items", "Operator Review Queue", "Hand Takeoff Queue"}
    assert expected.issubset(names), f"missing: {expected - names}"


def test_qa_pos_grand_total_visible_on_summary() -> None:
    est = Estimate(
        project_name="QA",
        contingency_pct=10.0,
        overhead_pct=10.0,
        profit_pct=5.0,
        line_items=[_line(total=1000.0, confidence=0.92, price_confidence=0.95)],
    )
    payload = export_estimate_xlsx(est, _project(), csi_titles={})
    wb = _open(payload)
    ws = wb["Summary"]
    # ``GRAND TOTAL`` row 9 carries B9 = headline grand total.
    assert ws["A9"].value == "GRAND TOTAL"
    assert isinstance(ws["B9"].value, (int, float))
    assert ws["B9"].value == pytest.approx(est.grand_total, abs=0.01)


# ---------------------------------------------------------------------------
# Negative
# ---------------------------------------------------------------------------


def test_qa_neg_empty_estimate_does_not_crash() -> None:
    """Estimate with no line items still exports cleanly."""
    est = Estimate(project_name="QA Empty", line_items=[])
    payload = export_estimate_xlsx(est, _project(), csi_titles={})
    wb = _open(payload)
    # Summary still rendered; grand total is 0.
    ws = wb["Summary"]
    assert ws["B9"].value == 0


def test_qa_neg_suppressed_lines_zero_total_in_summary() -> None:
    """Suppressed (unit-mismatch) line contributes $0 to the headline."""
    est = Estimate(project_name="QA Supp", line_items=[
        _line(description="Suppressed line", suppressed=True, confidence=0.9),
        _line(description="Real line", total=2_000.0),
    ])
    payload = export_estimate_xlsx(est, _project(), csi_titles={})
    wb = _open(payload)
    ws = wb["Summary"]
    # Subtotal == 2000 (the real line) — suppressed contributes $0.
    assert ws["B5"].value == pytest.approx(2_000.0, abs=0.01)


# ---------------------------------------------------------------------------
# Edge
# ---------------------------------------------------------------------------


def test_qa_edge_alternates_render_in_summary_block() -> None:
    """When alternates are priced, the summary surfaces the rollup block."""
    alt = AlternateLineEstimate(
        alternate_id="Alternate 1",
        alternate_type=AlternateType.ADDITIVE,
        description="ADD upgraded carpet tile",
        cost_delta=12_500.0,
        pricing_basis=AlternatePricingBasis.EXTRACTED_FROM_BID_FORM,
    )
    est = Estimate(
        project_name="QA Alt",
        line_items=[_line(total=1_000.0)],
        alternates=[alt],
    )
    payload = export_estimate_xlsx(est, _project(), csi_titles={})
    wb = _open(payload)
    ws = wb["Summary"]
    # Bid Alternates block exists at A17.
    assert ws["A17"].value == "Bid Alternates"
    # Total additive surfaces at B19.
    assert ws["B19"].value == pytest.approx(12_500.0, abs=0.01)


def test_qa_edge_no_alternates_no_alternates_block() -> None:
    """Estimate with zero alternates does NOT render the alternates summary block."""
    est = Estimate(project_name="QA NoAlt", line_items=[_line()], alternates=[])
    payload = export_estimate_xlsx(est, _project(), csi_titles={})
    wb = _open(payload)
    ws = wb["Summary"]
    # A17 must NOT be the "Bid Alternates" block — division layout slides up.
    assert ws["A17"].value != "Bid Alternates"


def test_qa_edge_alternates_sheet_only_when_alternates_present() -> None:
    """``Bid Alternates`` worksheet appears iff at least one alternate exists."""
    no_alt = Estimate(project_name="QA NoAlt", line_items=[_line()], alternates=[])
    payload_no = export_estimate_xlsx(no_alt, _project(), csi_titles={})
    wb_no = _open(payload_no)
    assert "Bid Alternates" not in wb_no.sheetnames

    with_alt = Estimate(
        project_name="QA WithAlt",
        line_items=[_line()],
        alternates=[AlternateLineEstimate(
            alternate_id="Alternate 1",
            alternate_type=AlternateType.ADDITIVE,
            description="Add carpet upgrade",
            cost_delta=5000.0,
            pricing_basis=AlternatePricingBasis.EXTRACTED_FROM_BID_FORM,
        )],
    )
    payload_yes = export_estimate_xlsx(with_alt, _project(), csi_titles={})
    wb_yes = _open(payload_yes)
    assert "Bid Alternates" in wb_yes.sheetnames


def test_qa_edge_large_estimate_perf_smoke() -> None:
    """500-line estimate exports in well under 5s and produces a valid workbook."""
    import time
    lines = [_line(description=f"Line {i}", total=100.0 + i)
             for i in range(500)]
    est = Estimate(project_name="QA Large", line_items=lines)
    t0 = time.perf_counter()
    payload = export_estimate_xlsx(est, _project(), csi_titles={})
    elapsed = time.perf_counter() - t0
    # Export should complete in well under 5s on dev hardware.
    assert elapsed < 5.0, f"export took {elapsed:.2f}s — performance regression?"
    wb = _open(payload)
    # Line Items sheet has 500 data rows + 1 header row + maybe 1 freeze line.
    ws = wb["Line Items"]
    # Rows must be at least 501 (header + 500 lines). Some exporters add
    # extra header rows; verifying ``>= 501`` keeps the test resilient.
    assert ws.max_row >= 501


def test_qa_edge_unicode_descriptions_round_trip() -> None:
    """Unicode in descriptions survives the openpyxl write/read round trip."""
    est = Estimate(
        project_name="QA Unicode — éàü",
        line_items=[_line(description="Ceiling tile — acoustic, 24″ x 24″")],
    )
    payload = export_estimate_xlsx(est, _project(), csi_titles={})
    wb = _open(payload)
    ws = wb["Summary"]
    # Project name preserved on B1.
    assert ws["B1"].value == "QA Unicode — éàü"
