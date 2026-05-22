"""Tests for the calibration-v3 Supporting Documents split.

Cross-cutting reference documents — wage determinations (Davis-Bacon,
TX prevailing wage), sample/template CSAs, tax-exemption certificates,
HSP form templates, UGSC / SGC, AIA contract templates — were previously
landing in the Bid Packages table with `trade_name='other'`. Calibration
v3 introduced a `document_kind` field on `BidPackage` and a deterministic
filename + first-page-text heuristic to classify them before the LLM
call.

These tests confirm:
  * `_supporting_doc_hint` recognizes the common variants
  * `ProjectModel.supporting_documents` / `trade_packages` filter cleanly
  * Excel export creates both sheets and they are disjoint
  * PDF export adds a "Supporting Documents" page when there are docs
    and skips the page entirely when there aren't (no empty section)

Filename heuristic fixtures are deliberately small string snippets — no
LLM calls, no real PDF parsing.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from core.exporter import export_estimate_xlsx
from core.exporter_pdf import build_quote_pdf
from core.extractors import _supporting_doc_hint
from core.schemas import (
    BidPackage,
    Estimate,
    PaymentMilestone,
    PaymentSchedule,
    QuoteConfig,
    QuoteMeta,
    SiteInfo,
)
from core.takeoff import ProjectInfo, ProjectModel, ScopeMatrix


# ---------------------------------------------------------------------------
# _supporting_doc_hint — filename + first-page text heuristic
# ---------------------------------------------------------------------------


def test_hint_recognises_davis_bacon_wage_determination() -> None:
    assert _supporting_doc_hint(
        filename="SAM_-_Davis-Bacon_Act_WD_TX20260254__Hays_County__Building.pdf",
        first_page_text="GENERAL DECISION TX20260254 — Davis-Bacon Act wage rates",
    )


def test_hint_recognises_prevailing_wage_table() -> None:
    assert _supporting_doc_hint(
        filename="Attachment_F.1_Tom_Green_County_Prevailing_Wage_2023.pdf",
        first_page_text="PREVAILING WAGE RATES TABLE — Tom Green County",
    )


def test_hint_recognises_tax_exemption_cert() -> None:
    assert _supporting_doc_hint(
        filename="Attachment E - Tax Exemption Certificate.pdf",
        first_page_text="Sales Tax Exemption Certificate for educational institution.",
    )


def test_hint_recognises_sample_csa() -> None:
    assert _supporting_doc_hint(
        filename="Attachment_B.2_Sample_Construction_Services_Agreement.pdf",
        first_page_text="SAMPLE CONSTRUCTION SERVICES AGREEMENT — template only",
    )


def test_hint_recognises_hsp_template() -> None:
    assert _supporting_doc_hint(
        filename="Attachment_D_HSP_HUB_Subcontracting_Plan.pdf",
        first_page_text="HUB Subcontracting Plan form to be completed by bidders.",
    )


def test_hint_recognises_ugsc() -> None:
    assert _supporting_doc_hint(
        filename="Attachment_C_2010_Uniform_General_Conditions.pdf",
        first_page_text="UGSC — Uniform General Conditions for State Construction.",
    )


def test_hint_returns_false_for_normal_trade_package() -> None:
    """A vanilla trade package should NOT trip the heuristic."""
    assert not _supporting_doc_hint(
        filename="Beck_03.00_Turnkey_Structural_Concrete.pdf",
        first_page_text="BID PACKAGE 03.00 — STRUCTURAL CONCRETE. Specific inclusions...",
    )


def test_hint_handles_empty_inputs() -> None:
    assert not _supporting_doc_hint("", "")
    assert not _supporting_doc_hint(None, None)  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# ProjectModel.supporting_documents / trade_packages properties
# ---------------------------------------------------------------------------


def _mixed_project() -> ProjectModel:
    pkgs = [
        BidPackage(
            pdf_name="trade_concrete.pdf",
            package_number="03",
            trade_name="Concrete",
            document_kind="trade_package",
        ),
        BidPackage(
            pdf_name="trade_drywall.pdf",
            package_number="09",
            trade_name="Drywall",
            document_kind="trade_package",
        ),
        BidPackage(
            pdf_name="wage_determination_TX20260254.pdf",
            trade_name=None,
            document_kind="supporting_document",
            summary="Davis-Bacon wage determination, Hays County, effective Jan 1, 2026.",
        ),
        BidPackage(
            pdf_name="sample_csa_template.pdf",
            trade_name=None,
            document_kind="supporting_document",
            summary="Sample construction services agreement template.",
        ),
    ]
    return ProjectModel(
        rooms=[], doors=[], windows=[], structural=[], mep=[], spec_sections=[],
        site=SiteInfo(),
        takeoffs=[], sheet_summaries={}, warnings=[],
        bid_packages=pkgs,
        project_info=ProjectInfo(name="Mixed", number="0"),
        scope_matrix=ScopeMatrix(packages=pkgs, by_division={}, all_alternates=[], coverage_warnings=[]),
        aggregated_inclusions=[], aggregated_exclusions=[],
    )


def test_supporting_documents_property_filters_correctly() -> None:
    project = _mixed_project()
    # Single source of truth: master list keeps both kinds.
    assert len(project.bid_packages) == 4
    # Filtered views are disjoint and complete.
    trades = project.trade_packages
    supps = project.supporting_documents
    assert {p.pdf_name for p in trades} == {"trade_concrete.pdf", "trade_drywall.pdf"}
    assert {p.pdf_name for p in supps} == {
        "wage_determination_TX20260254.pdf",
        "sample_csa_template.pdf",
    }
    assert len(trades) + len(supps) == len(project.bid_packages)


# ---------------------------------------------------------------------------
# Excel export — Bid Packages and Supporting Documents are disjoint
# ---------------------------------------------------------------------------


def test_excel_export_creates_both_sheets_disjoint() -> None:
    project = _mixed_project()
    estimate = Estimate(project_name="Mixed")
    payload = export_estimate_xlsx(
        estimate=estimate, project=project, csi_titles={}, extractions=[],
    )
    wb = load_workbook(BytesIO(payload))
    assert "Bid Packages" in wb.sheetnames
    assert "Supporting Documents" in wb.sheetnames

    # Bid Packages must NOT contain any supporting doc filenames.
    bp_ws = wb["Bid Packages"]
    bp_headers = [c.value for c in bp_ws[1]]
    pdf_col = bp_headers.index("PDF") + 1
    bp_pdfs = {bp_ws.cell(row=r, column=pdf_col).value for r in range(2, bp_ws.max_row + 1)}
    assert bp_pdfs == {"trade_concrete.pdf", "trade_drywall.pdf"}
    assert "wage_determination_TX20260254.pdf" not in bp_pdfs
    assert "sample_csa_template.pdf" not in bp_pdfs

    # Supporting Documents must contain ONLY supporting docs.
    sd_ws = wb["Supporting Documents"]
    sd_headers = [c.value for c in sd_ws[1]]
    assert "Filename" in sd_headers
    assert "Kind" in sd_headers
    fn_col = sd_headers.index("Filename") + 1
    kind_col = sd_headers.index("Kind") + 1
    sd_pdfs = {sd_ws.cell(row=r, column=fn_col).value for r in range(2, sd_ws.max_row + 1)}
    assert sd_pdfs == {
        "wage_determination_TX20260254.pdf",
        "sample_csa_template.pdf",
    }
    sd_kinds = {sd_ws.cell(row=r, column=kind_col).value for r in range(2, sd_ws.max_row + 1)}
    assert "wage determination" in sd_kinds
    assert "sample agreement" in sd_kinds


def test_excel_export_omits_supporting_documents_sheet_when_empty() -> None:
    """No supporting docs => no Supporting Documents sheet (clean export)."""
    pkgs = [BidPackage(pdf_name="t.pdf", trade_name="x", document_kind="trade_package")]
    project = ProjectModel(
        rooms=[], doors=[], windows=[], structural=[], mep=[], spec_sections=[],
        site=SiteInfo(), takeoffs=[], sheet_summaries={}, warnings=[],
        bid_packages=pkgs,
        project_info=ProjectInfo(name="x", number="0"),
        scope_matrix=ScopeMatrix(packages=pkgs, by_division={}, all_alternates=[], coverage_warnings=[]),
        aggregated_inclusions=[], aggregated_exclusions=[],
    )
    payload = export_estimate_xlsx(
        estimate=Estimate(project_name="x"),
        project=project, csi_titles={}, extractions=[],
    )
    wb = load_workbook(BytesIO(payload))
    assert "Supporting Documents" not in wb.sheetnames


# ---------------------------------------------------------------------------
# PDF export — Supporting Documents page is conditional
# ---------------------------------------------------------------------------


def _default_quote_config() -> QuoteConfig:
    return QuoteConfig(
        quote_meta=QuoteMeta(scope_blurb="x"),
        payment_schedule=PaymentSchedule(
            mode="percentage",
            milestones=[PaymentMilestone(label="Done", percentage=100.0)],
        ),
    )


def _priced_estimate() -> Estimate:
    """Tiny non-empty estimate so the executive summary renders the normal
    three-tile path (not the empty-state banner — that's covered in
    test_pdf_empty_state.py)."""
    from core.schemas import CostCategory, CostLine
    return Estimate(
        project_name="Mixed",
        line_items=[
            CostLine(
                csi_division="03", csi_section="03 30 00",
                description="Concrete", quantity=100.0, unit="SF",
                unit_cost=10.0, total_cost=1000.0,
                cost_category=CostCategory.MATERIAL,
            ),
        ],
    )


def _pdf_text(pdf_path: Path) -> str:
    """Extract text from every page via PyMuPDF (already a project dep)."""
    import fitz  # PyMuPDF
    with fitz.open(pdf_path) as doc:
        return "\n".join(page.get_text() for page in doc)


def test_pdf_export_supporting_documents_page_present_when_docs_exist(tmp_path: Path) -> None:
    out = tmp_path / "with_supp_docs.pdf"
    build_quote_pdf(
        estimate=_priced_estimate(),
        project=_mixed_project(),
        quote_config=_default_quote_config(),
        out_path=out,
    )
    assert out.is_file()
    assert out.read_bytes()[:5] == b"%PDF-"

    text = _pdf_text(out)
    assert "Supporting documents" in text, \
        "Supporting documents heading missing from PDF"
    # The filenames of the two supporting docs should appear in the page.
    assert "wage_determination_TX20260254.pdf" in text
    assert "sample_csa_template.pdf" in text


def test_pdf_export_skips_supporting_documents_page_when_empty(tmp_path: Path) -> None:
    """Empty supporting-docs list => no page rendered, no empty section."""
    pkgs = [BidPackage(pdf_name="t.pdf", trade_name="x", document_kind="trade_package")]
    project = ProjectModel(
        rooms=[], doors=[], windows=[], structural=[], mep=[], spec_sections=[],
        site=SiteInfo(), takeoffs=[], sheet_summaries={}, warnings=[],
        bid_packages=pkgs,
        project_info=ProjectInfo(name="x", number="0"),
        scope_matrix=ScopeMatrix(packages=pkgs, by_division={}, all_alternates=[], coverage_warnings=[]),
        aggregated_inclusions=[], aggregated_exclusions=[],
    )
    out = tmp_path / "no_supp_docs.pdf"
    build_quote_pdf(
        estimate=_priced_estimate(),
        project=project,
        quote_config=_default_quote_config(),
        out_path=out,
    )
    assert out.read_bytes()[:5] == b"%PDF-"
    text = _pdf_text(out)
    assert "Supporting documents" not in text, \
        "PDF should not render the Supporting documents heading when the list is empty"
