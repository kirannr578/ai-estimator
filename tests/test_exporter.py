"""Phase T6 — Excel exporter band-aware tests.

Covers:

* Operator Review Queue sheet appears (even when empty).
* Hand Takeoff Queue sheet appears (even when empty).
* Hand Takeoff Queue header is tinted with ``HAND_TAKEOFF_HEADER_FILL``.
* Operator Review Queue header is tinted with ``OPERATOR_REVIEW_HEADER_FILL``.
* The "Line Items" sheet carries a ``Band`` column with values from
  ``{AUTO, REVIEW, HAND}``.
* The Project Summary sheet carries the seven new T6 rows.
* Queue sheets carry only the rows that belong to their band; the
  AUTO-band line never leaks into either queue.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.exporter import (
    HAND_TAKEOFF_HEADER_FILL,
    OPERATOR_REVIEW_HEADER_FILL,
    export_estimate_xlsx,
)
from core.schemas import (
    CostBand,
    CostCategory,
    CostLine,
    Estimate,
    SiteInfo,
    band_for_confidence,
)
from core.takeoff import ProjectInfo, ProjectModel, ScopeMatrix


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _empty_project() -> ProjectModel:
    return ProjectModel(
        rooms=[],
        doors=[],
        windows=[],
        structural=[],
        mep=[],
        spec_sections=[],
        site=SiteInfo(),
        takeoffs=[],
        sheet_summaries={},
        warnings=[],
        bid_packages=[],
        project_info=ProjectInfo(name="T6 Export Test"),
        scope_matrix=ScopeMatrix(
            packages=[], by_division={}, all_alternates=[], coverage_warnings=[]
        ),
        aggregated_inclusions=[],
        aggregated_exclusions=[],
    )


def _line(
    *,
    division: str = "09",
    section: str | None = "09 91 23",
    description: str = "Interior painting",
    confidence: float = 0.92,
    total: float = 1000.0,
    suppressed: bool = False,
) -> CostLine:
    return CostLine(
        csi_division=division,
        csi_section=section,
        description=description,
        quantity=500.0,
        unit="SF",
        unit_cost=round(total / 500.0, 4),
        total_cost=total,
        cost_category=CostCategory.SUBCONTRACTOR,
        confidence=confidence,
        suppressed=suppressed,
        cost_band=band_for_confidence(confidence, suppressed=suppressed),
    )


def _mixed_estimate() -> Estimate:
    """Three priced lines (one per band) + one suppressed (HAND)."""
    return Estimate(
        project_name="T6 Export",
        region_multiplier=1.0,
        contingency_pct=10.0,
        overhead_pct=10.0,
        profit_pct=5.0,
        line_items=[
            _line(description="Auto", confidence=0.92, total=10_000.0),
            _line(description="Review", confidence=0.78, total=5_000.0),
            _line(description="Hand", confidence=0.40, total=2_000.0),
            _line(
                description="Mismatch suppressed",
                confidence=0.99, total=0.0, suppressed=True,
            ),
        ],
    )


def _all_auto_estimate() -> Estimate:
    return Estimate(
        project_name="all-auto",
        line_items=[
            _line(description="A", confidence=0.92, total=1000.0),
            _line(description="B", confidence=0.99, total=2000.0),
        ],
    )


def _load_xlsx(payload: bytes):
    return load_workbook(BytesIO(payload))


# ---------------------------------------------------------------------------
# Sheet presence
# ---------------------------------------------------------------------------


def test_operator_review_queue_sheet_appears() -> None:
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    assert "Operator Review Queue" in wb.sheetnames


def test_hand_takeoff_queue_sheet_appears() -> None:
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    assert "Hand Takeoff Queue" in wb.sheetnames


def test_queue_sheets_appear_even_when_empty() -> None:
    """All-AUTO estimate: both queue sheets must still exist (no crash)."""
    payload = export_estimate_xlsx(_all_auto_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    assert "Operator Review Queue" in wb.sheetnames
    assert "Hand Takeoff Queue" in wb.sheetnames


# ---------------------------------------------------------------------------
# Header tinting
# ---------------------------------------------------------------------------


def _header_fill_rgb(ws) -> str | None:
    cell = ws.cell(row=1, column=1)
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return None
    rgb = fill.fgColor.rgb
    if rgb is None:
        return None
    return str(rgb).upper()


def test_hand_takeoff_queue_header_uses_amber_tint() -> None:
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Hand Takeoff Queue"]
    # openpyxl serialises HAND fill as `00FFC107` (alpha + RGB) — match
    # by suffix so we don't depend on the alpha channel.
    rgb = _header_fill_rgb(ws)
    assert rgb is not None
    assert rgb.endswith(HAND_TAKEOFF_HEADER_FILL.fgColor.rgb.upper()), rgb


def test_operator_review_queue_header_uses_yellow_tint() -> None:
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Operator Review Queue"]
    rgb = _header_fill_rgb(ws)
    assert rgb is not None
    assert rgb.endswith(OPERATOR_REVIEW_HEADER_FILL.fgColor.rgb.upper()), rgb


# ---------------------------------------------------------------------------
# Band column on the "Line Items" sheet
# ---------------------------------------------------------------------------


def test_line_items_sheet_has_band_column() -> None:
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Line Items"]
    headers = [c.value for c in ws[1]]
    assert "Band" in headers, f"Expected Band column; got {headers}"


def test_line_items_band_values_are_in_short_label_set() -> None:
    """Band column values must be one of AUTO / REVIEW / HAND."""
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Line Items"]
    headers = [c.value for c in ws[1]]
    band_col = headers.index("Band") + 1
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=band_col).value
        # Skip division-banner rows where the Band column is blank.
        if v is None or v == "":
            continue
        seen.add(str(v))
    # Mixed estimate has one of each band.
    assert seen == {"AUTO", "REVIEW", "HAND"}, seen


# ---------------------------------------------------------------------------
# Queue-row content
# ---------------------------------------------------------------------------


def test_operator_review_queue_contains_only_review_rows() -> None:
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Operator Review Queue"]
    # Description column is index 2 in the queue schema.
    descs = [
        ws.cell(row=r, column=2).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=2).value
    ]
    assert "Review" in descs
    assert "Auto" not in descs
    assert "Hand" not in descs
    assert "Mismatch suppressed" not in descs


def test_hand_takeoff_queue_contains_low_conf_and_suppressed_rows() -> None:
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Hand Takeoff Queue"]
    descs = [
        ws.cell(row=r, column=2).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=2).value
    ]
    assert "Hand" in descs
    assert "Mismatch suppressed" in descs
    assert "Auto" not in descs
    assert "Review" not in descs


def test_empty_queue_sheet_has_no_data_rows_just_friendly_note() -> None:
    payload = export_estimate_xlsx(_all_auto_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Hand Takeoff Queue"]
    # Headers in row 1; friendly note in row 2 column 1; no data beyond.
    note = ws.cell(row=2, column=1).value
    assert note is not None and "No hand-takeoff lines" in str(note)


# ---------------------------------------------------------------------------
# Project Summary band rows
# ---------------------------------------------------------------------------


def test_summary_sheet_has_seven_new_t6_rows() -> None:
    """The Summary sheet must carry the seven brief-mandated T6 rows."""
    payload = export_estimate_xlsx(_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Summary"]
    # Collect column A labels from rows 10–20 (T6 rows live in 10..16).
    labels = []
    for r in range(10, 18):
        v = ws.cell(row=r, column=1).value
        if v:
            labels.append(str(v))

    expected_phrases = [
        "Auto-Approve Total",
        "Operator-Review Total",
        "Hand-Takeoff Total",
        "Grand Total (Auto-Only)",
        "Grand Total (Auto + Review)",
        "Lines Needing Manual Takeoff",
        "Lines Needing Operator Review",
    ]
    for phrase in expected_phrases:
        assert any(phrase in l for l in labels), (
            f"Summary sheet missing T6 row '{phrase}'; got {labels}"
        )


def test_summary_sheet_t6_dollar_values_match_estimate() -> None:
    est = _mixed_estimate()
    payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Summary"]

    # Row 10: Auto-Approve Total
    assert ws["A10"].value == "Auto-Approve Total"
    assert ws["B10"].value == pytest.approx(est.total_auto_approve)
    # Row 11: Operator-Review Total
    assert ws["B11"].value == pytest.approx(est.total_operator_review)
    # Row 12: Hand-Takeoff Total
    assert ws["B12"].value == pytest.approx(est.total_hand_takeoff)
    # Row 13: Grand Total (Auto-Only)
    assert ws["B13"].value == pytest.approx(est.grand_total_auto_only)
    # Row 14: Grand Total (Auto + Review)
    assert ws["B14"].value == pytest.approx(est.grand_total_with_review)


# ---------------------------------------------------------------------------
# JSON export band aggregates
# ---------------------------------------------------------------------------


def test_export_json_carries_band_aggregates() -> None:
    """JSON export must surface T6 fields for downstream consumers."""
    import json
    from core.exporter import export_estimate_json

    est = _mixed_estimate()
    payload = json.loads(export_estimate_json(est, _empty_project()))
    assert payload["grand_total_with_review"] == pytest.approx(est.grand_total_with_review)
    assert payload["grand_total_auto_only"] == pytest.approx(est.grand_total_auto_only)
    assert payload["total_auto_approve"] == pytest.approx(est.total_auto_approve)
    assert payload["total_operator_review"] == pytest.approx(est.total_operator_review)
    assert payload["total_hand_takeoff"] == pytest.approx(est.total_hand_takeoff)
    assert payload["auto_approve_count"] == est.auto_approve_count
    assert payload["operator_review_count"] == est.operator_review_count
    assert payload["hand_takeoff_count"] == est.hand_takeoff_count
    # Each serialised CostLine must carry its cost_band string.
    for li in payload["line_items"]:
        assert li["cost_band"] in {b.value for b in CostBand}


# ---------------------------------------------------------------------------
# Phase T7 — price_confidence + cost_source_tier + tier breakdown
# ---------------------------------------------------------------------------
#
# These tests pin the three new Cost Estimate columns, the Project
# Summary "Cost Source Tier Breakdown" block, and the all-EXACT /
# all-MISSING edge cases.


def _t7_line(
    *,
    division: str = "09",
    section: str | None = "09 91 23",
    description: str = "T7 line",
    confidence: float = 0.95,
    price_confidence: float = 1.0,
    tier=None,
    total: float = 1000.0,
    suppressed: bool = False,
) -> CostLine:
    from core.schemas import CostSourceTier as _Tier
    return CostLine(
        csi_division=division,
        csi_section=section,
        description=description,
        quantity=500.0,
        unit="SF",
        unit_cost=round(total / 500.0, 4),
        total_cost=total,
        cost_category=CostCategory.SUBCONTRACTOR,
        confidence=confidence,
        price_confidence=price_confidence,
        cost_source_tier=tier if tier is not None else _Tier.EXACT_MATCH,
        suppressed=suppressed,
        cost_band=band_for_confidence(
            round(confidence * price_confidence, 4),
            suppressed=suppressed,
        ),
    )


def _t7_mixed_estimate() -> Estimate:
    """One line per Phase T7 tier, totals chosen to make the percentage
    arithmetic obvious (rounds to whole percents at $100 / $200 / $300)."""
    from core.schemas import CostSourceTier as _Tier
    return Estimate(
        project_name="T7 Export",
        region_multiplier=1.0,
        contingency_pct=10.0,
        overhead_pct=10.0,
        profit_pct=5.0,
        line_items=[
            _t7_line(
                description="exact",
                confidence=0.95, price_confidence=0.95,
                tier=_Tier.EXACT_MATCH, total=1000.0,
            ),
            _t7_line(
                description="category",
                confidence=0.92, price_confidence=0.78,
                tier=_Tier.CATEGORY_MATCH, total=500.0,
            ),
            _t7_line(
                description="interpolated",
                confidence=0.92, price_confidence=0.65,
                tier=_Tier.INTERPOLATED, total=300.0,
            ),
            _t7_line(
                description="parametric",
                confidence=0.92, price_confidence=0.45,
                tier=_Tier.PARAMETRIC, total=200.0,
            ),
            _t7_line(
                description="missing",
                confidence=0.92, price_confidence=0.0,
                tier=_Tier.MISSING, total=0.0, suppressed=True,
            ),
        ],
    )


def _t7_all_exact_estimate() -> Estimate:
    from core.schemas import CostSourceTier as _Tier
    return Estimate(
        project_name="all-exact",
        line_items=[
            _t7_line(
                description="A",
                confidence=0.95, price_confidence=0.95,
                tier=_Tier.EXACT_MATCH, total=1000.0,
            ),
            _t7_line(
                description="B",
                confidence=0.92, price_confidence=0.95,
                tier=_Tier.EXACT_MATCH, total=2000.0,
            ),
        ],
    )


def _t7_all_missing_estimate() -> Estimate:
    from core.schemas import CostSourceTier as _Tier
    return Estimate(
        project_name="all-missing",
        line_items=[
            _t7_line(
                description="m1",
                confidence=0.92, price_confidence=0.0,
                tier=_Tier.MISSING, total=0.0, suppressed=True,
            ),
            _t7_line(
                description="m2",
                confidence=0.92, price_confidence=0.0,
                tier=_Tier.MISSING, total=0.0, suppressed=True,
            ),
        ],
    )


def test_t7_line_items_sheet_has_three_new_columns_in_order() -> None:
    """Cost Estimate sheet must carry Price Confidence, Cost Source Tier,
    and Combined Confidence — in that order — immediately after the
    legacy Confidence column."""
    payload = export_estimate_xlsx(_t7_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Line Items"]
    headers = [c.value for c in ws[1]]
    assert "Confidence" in headers
    assert "Price Confidence" in headers
    assert "Cost Source Tier" in headers
    assert "Combined Confidence" in headers
    conf_idx = headers.index("Confidence")
    pc_idx = headers.index("Price Confidence")
    tier_idx = headers.index("Cost Source Tier")
    comb_idx = headers.index("Combined Confidence")
    assert pc_idx == conf_idx + 1
    assert tier_idx == conf_idx + 2
    assert comb_idx == conf_idx + 3


def test_t7_tier_values_displayed_as_title_case_strings() -> None:
    """Cost Source Tier cells render as Title Case (matches `_TIER_LABELS`),
    not the enum string value."""
    from core.exporter import _TIER_LABELS
    from core.schemas import CostSourceTier as _Tier

    payload = export_estimate_xlsx(_t7_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Line Items"]
    headers = [c.value for c in ws[1]]
    tier_col = headers.index("Cost Source Tier") + 1
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=tier_col).value
        if v is None or v == "":
            continue
        seen.add(str(v))
    expected = {
        _TIER_LABELS[_Tier.EXACT_MATCH],
        _TIER_LABELS[_Tier.CATEGORY_MATCH],
        _TIER_LABELS[_Tier.INTERPOLATED],
        _TIER_LABELS[_Tier.PARAMETRIC],
        _TIER_LABELS[_Tier.MISSING],
    }
    assert seen >= expected, f"missing tiers: {expected - seen}"
    # Title case sanity — every label's first letter is upper case.
    for label in seen:
        assert label[0].isupper(), f"tier label not Title Case: {label!r}"


def test_t7_summary_has_cost_source_tier_breakdown_section() -> None:
    """Project Summary sheet must carry a 'Cost Source Tier Breakdown'
    header row with the per-tier breakdown table beneath it."""
    payload = export_estimate_xlsx(_t7_mixed_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Summary"]
    found = False
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Cost Source Tier Breakdown":
            found = True
            break
    assert found, "Project Summary missing 'Cost Source Tier Breakdown' header"


def test_t7_summary_tier_breakdown_carries_count_dollars_and_percent() -> None:
    """Per-tier rows must include count + $ + % of subtotal columns."""
    from core.exporter import _TIER_LABELS
    from core.schemas import CostSourceTier as _Tier

    est = _t7_mixed_estimate()
    payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Summary"]

    # Find the breakdown header row, then walk down to read the table rows.
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Cost Source Tier Breakdown":
            header_row = r
            break
    assert header_row is not None
    # The Tier / Line Count / Total $ / % of Subtotal header row sits
    # immediately below the section header.
    table_header_row = header_row + 1
    table_headers = [
        ws.cell(row=table_header_row, column=c).value for c in range(1, 5)
    ]
    assert table_headers == ["Tier", "Line Count", "Total $", "% of Subtotal"]

    # Walk forward and collect tier-row data until we hit the totals
    # bold-row (which starts with "Total: N lines").
    rows: dict[str, tuple[int, float]] = {}
    for r in range(table_header_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label is None or str(label).startswith("Total:"):
            break
        cnt = ws.cell(row=r, column=2).value
        amt = ws.cell(row=r, column=3).value
        rows[str(label)] = (int(cnt or 0), float(amt or 0.0))

    expected_counts = est.count_by_tier
    expected_totals = est.total_by_tier
    for tier, label in _TIER_LABELS.items():
        cnt, amt = rows.get(label, (0, 0.0))
        assert cnt == expected_counts[tier], f"{label} count"
        assert amt == pytest.approx(expected_totals[tier]), f"{label} total"


def test_t7_summary_tier_breakdown_appears_for_all_exact_estimate() -> None:
    """All-EXACT estimate still emits the breakdown block (other tiers
    show zeros) so downstream parsers can rely on it always existing."""
    payload = export_estimate_xlsx(_t7_all_exact_estimate(), _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Summary"]
    found = False
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Cost Source Tier Breakdown":
            found = True
            break
    assert found


def test_t7_summary_tier_breakdown_for_all_missing_estimate_has_zero_dollars() -> None:
    """All-MISSING estimate (every line suppressed at $0): breakdown
    appears with the MISSING row showing a non-zero count but $0 total."""
    from core.exporter import _TIER_LABELS
    from core.schemas import CostSourceTier as _Tier

    est = _t7_all_missing_estimate()
    payload = export_estimate_xlsx(est, _empty_project(), csi_titles={})
    wb = _load_xlsx(payload)
    ws = wb["Summary"]

    # Subtotal must be 0 (every line is suppressed → HAND_TAKEOFF →
    # excluded from headline).
    assert est.subtotal == 0.0

    # Find the MISSING row in the breakdown.
    missing_label = _TIER_LABELS[_Tier.MISSING]
    missing_count = None
    missing_total = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == missing_label:
            missing_count = ws.cell(row=r, column=2).value
            missing_total = ws.cell(row=r, column=3).value
            break
    assert missing_count == 2
    assert missing_total == pytest.approx(0.0)


def test_t7_export_json_carries_tier_aggregates() -> None:
    """JSON export must surface the Phase T7 tier aggregates alongside
    the T6 band aggregates."""
    import json
    from core.exporter import export_estimate_json
    from core.schemas import CostSourceTier as _Tier

    est = _t7_mixed_estimate()
    payload = json.loads(export_estimate_json(est, _empty_project()))
    assert "total_by_tier" in payload
    assert "count_by_tier" in payload
    # All six tier keys are present using the enum string values.
    for tier in _Tier:
        assert tier.value in payload["total_by_tier"]
        assert tier.value in payload["count_by_tier"]
    # Round-trip the EXACT_MATCH counts + dollars.
    assert payload["count_by_tier"][_Tier.EXACT_MATCH.value] == (
        est.count_by_tier[_Tier.EXACT_MATCH]
    )
    assert payload["total_by_tier"][_Tier.EXACT_MATCH.value] == pytest.approx(
        est.total_by_tier[_Tier.EXACT_MATCH]
    )
    # Each line carries the new T7 fields.
    for li in payload["line_items"]:
        assert "price_confidence" in li
        assert "cost_source_tier" in li
        assert li["cost_source_tier"] in {t.value for t in _Tier}
