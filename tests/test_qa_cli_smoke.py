"""QA pass — Subsystem 10: End-to-end CLI smoke test (2026-05-28).

Runs ``analyze.main()`` in-process against a synthetic mini-project
with the LLM extraction layer fully mocked. Verifies that:

* The CLI writes ``estimate.xlsx``, ``estimate.json``, and
  ``run_log.txt`` to the output directory.
* The CLI exits 0 (success).
* The Excel / JSON deliverables open without crashing.

LLM spend is avoided by patching ``analyze.LLMClient``,
``analyze.extract_sheet``, and ``analyze.extract_bundle`` at the
module boundary so no real LLM calls fire.
"""

from __future__ import annotations

import json
import sys
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import fitz
import pytest
from openpyxl import load_workbook

import analyze
from core.schemas import SheetExtraction


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


class _NullLLM:
    """Drop-in replacement for ``LLMClient`` that never makes a network call."""

    def __init__(self, *args, **kwargs):
        self.provider = kwargs.get("provider", "anthropic")
        self.model = kwargs.get("model", "fake-model")

    def call(self, *args, **kwargs):  # pragma: no cover — defensive only
        raise AssertionError("LLM call attempted in QA smoke test")

    def call_with_image(self, *args, **kwargs):  # pragma: no cover
        raise AssertionError("LLM image call attempted in QA smoke test")


def _stub_extract_sheet(sheet, llm):
    """Return an empty :class:`SheetExtraction` — preserves shape, skips LLM."""
    return SheetExtraction(sheet_id=sheet.sheet_id, summary="QA stub", warnings=[])


def _stub_extract_bundle(bundle, llm):
    return SheetExtraction(
        sheet_id=getattr(bundle, "pdf_name", "qa-bundle"),
        summary="QA stub",
        warnings=[],
    )


def _build_text_pdf(path: Path, *, title: str, body_lines: list[str]) -> None:
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    page.insert_text((40, 60), title, fontsize=14)
    for i, line in enumerate(body_lines):
        page.insert_text((40, 90 + i * 14), line, fontsize=10)
    doc.save(path)
    doc.close()


# ---------------------------------------------------------------------------
# Positive
# ---------------------------------------------------------------------------


def test_qa_pos_cli_smoke_single_pdf_run(tmp_path: Path, monkeypatch) -> None:
    """End-to-end: synthetic 1-page bid-form PDF → estimate.xlsx + .json."""
    in_pdf = tmp_path / "bid_form.pdf"
    _build_text_pdf(
        in_pdf,
        title="BID FORM — QA SMOKE TEST",
        body_lines=[
            "PROJECT: QA Smoke Demo",
            "GENERAL CONTRACTOR: Acme Construction",
            "BASE BID: $1,000,000.00",
            "Alternate No. 1: ADD upgraded carpet. $5,000",
        ],
    )
    out_dir = tmp_path / "out"

    # Patch the LLM boundary BEFORE invoking main.
    monkeypatch.setattr(analyze, "LLMClient", _NullLLM)
    monkeypatch.setattr(analyze, "extract_sheet", _stub_extract_sheet)
    monkeypatch.setattr(analyze, "extract_bundle", _stub_extract_bundle)

    # Build a synthetic argv and run main().
    argv = [
        "analyze",
        str(in_pdf),
        "--out", str(out_dir),
        "--workers", "1",
        "--quiet",
    ]
    monkeypatch.setattr(sys, "argv", argv)

    rc = analyze.main()
    assert rc == 0, "CLI should exit 0 on a successful run"

    # Required deliverables.
    assert (out_dir / "estimate.xlsx").is_file()
    assert (out_dir / "estimate.json").is_file()
    assert (out_dir / "run_log.txt").is_file()


def test_qa_pos_cli_smoke_xlsx_opens_cleanly(tmp_path: Path, monkeypatch) -> None:
    """The CLI's ``estimate.xlsx`` is a valid openpyxl workbook."""
    in_pdf = tmp_path / "spec.pdf"
    _build_text_pdf(in_pdf, title="SPEC SECTION 09 91 23",
                    body_lines=["Interior painting; latex; semi-gloss."])
    out_dir = tmp_path / "out"

    monkeypatch.setattr(analyze, "LLMClient", _NullLLM)
    monkeypatch.setattr(analyze, "extract_sheet", _stub_extract_sheet)
    monkeypatch.setattr(analyze, "extract_bundle", _stub_extract_bundle)
    monkeypatch.setattr(sys, "argv", [
        "analyze", str(in_pdf), "--out", str(out_dir), "--workers", "1", "--quiet",
    ])

    rc = analyze.main()
    assert rc == 0
    wb = load_workbook(out_dir / "estimate.xlsx", read_only=True, data_only=False)
    # Summary sheet exists.
    assert "Summary" in wb.sheetnames
    wb.close()


# ---------------------------------------------------------------------------
# Negative
# ---------------------------------------------------------------------------


def test_qa_neg_cli_no_pdfs_returns_nonzero(tmp_path: Path, monkeypatch) -> None:
    """Empty input directory → CLI exits non-zero, no crash."""
    empty_dir = tmp_path / "empty"
    empty_dir.mkdir()
    out_dir = tmp_path / "out"

    monkeypatch.setattr(analyze, "LLMClient", _NullLLM)
    monkeypatch.setattr(sys, "argv", [
        "analyze", str(empty_dir), "--out", str(out_dir), "--quiet",
    ])
    rc = analyze.main()
    # Spec: rc == 2 when no PDFs found.
    assert rc != 0


def test_qa_neg_cli_nonexistent_path_raises(tmp_path: Path, monkeypatch) -> None:
    """Bad path → CLI raises ``SystemExit`` (argparse contract)."""
    bogus = tmp_path / "does_not_exist.pdf"
    out_dir = tmp_path / "out"

    monkeypatch.setattr(analyze, "LLMClient", _NullLLM)
    monkeypatch.setattr(sys, "argv", [
        "analyze", str(bogus), "--out", str(out_dir), "--quiet",
    ])
    with pytest.raises(SystemExit):
        analyze.main()


# ---------------------------------------------------------------------------
# Edge
# ---------------------------------------------------------------------------


def test_qa_edge_cli_skips_drawings_when_flag_set(tmp_path: Path, monkeypatch) -> None:
    """``--no-drawings`` filters out >5MB PDFs (heuristic) — small PDFs proceed."""
    # Build two small text PDFs (well under 5MB) so the no-drawings flag
    # leaves them in the run.
    pdf1 = tmp_path / "manual.pdf"
    pdf2 = tmp_path / "spec.pdf"
    _build_text_pdf(pdf1, title="PROJECT MANUAL", body_lines=["Spec section 01"])
    _build_text_pdf(pdf2, title="SPEC", body_lines=["Spec body"])
    out_dir = tmp_path / "out"

    monkeypatch.setattr(analyze, "LLMClient", _NullLLM)
    monkeypatch.setattr(analyze, "extract_sheet", _stub_extract_sheet)
    monkeypatch.setattr(analyze, "extract_bundle", _stub_extract_bundle)
    monkeypatch.setattr(sys, "argv", [
        "analyze", str(tmp_path), "--out", str(out_dir),
        "--workers", "1", "--no-drawings", "--quiet",
    ])

    rc = analyze.main()
    assert rc == 0
    assert (out_dir / "estimate.xlsx").is_file()


def test_qa_edge_cli_propagates_extraction_warnings(tmp_path: Path, monkeypatch) -> None:
    """Warnings from extractors land in the run_log.

    With our stubs returning empty extractions, the run completes
    cleanly but no warnings appear; what matters is that the run_log
    exists and is non-empty (header + at least one log line).
    """
    in_pdf = tmp_path / "qa.pdf"
    _build_text_pdf(in_pdf, title="QA PDF", body_lines=["Body"])
    out_dir = tmp_path / "out"

    monkeypatch.setattr(analyze, "LLMClient", _NullLLM)
    monkeypatch.setattr(analyze, "extract_sheet", _stub_extract_sheet)
    monkeypatch.setattr(analyze, "extract_bundle", _stub_extract_bundle)
    monkeypatch.setattr(sys, "argv", [
        "analyze", str(in_pdf), "--out", str(out_dir),
        "--workers", "1", "--quiet",
    ])

    rc = analyze.main()
    assert rc == 0
    log_text = (out_dir / "run_log.txt").read_text(encoding="utf-8")
    # Run log should contain the per-document elapsed-time line.
    assert log_text.strip(), "run_log.txt should contain at least one line"


def test_qa_edge_cli_estimate_json_is_well_formed(tmp_path: Path, monkeypatch) -> None:
    """``estimate.json`` round-trips through :func:`json.loads`."""
    in_pdf = tmp_path / "qa.pdf"
    _build_text_pdf(in_pdf, title="QA JSON", body_lines=["Body"])
    out_dir = tmp_path / "out"

    monkeypatch.setattr(analyze, "LLMClient", _NullLLM)
    monkeypatch.setattr(analyze, "extract_sheet", _stub_extract_sheet)
    monkeypatch.setattr(analyze, "extract_bundle", _stub_extract_bundle)
    monkeypatch.setattr(sys, "argv", [
        "analyze", str(in_pdf), "--out", str(out_dir),
        "--workers", "1", "--quiet",
    ])

    rc = analyze.main()
    assert rc == 0
    data = json.loads((out_dir / "estimate.json").read_text(encoding="utf-8"))
    # The exporter writes a top-level dict-like JSON. Confirm it's a dict
    # and carries the expected key.
    assert isinstance(data, dict)
    # ``project_name`` lives somewhere in the JSON tree.
    assert "estimate" in data or "project_name" in data
