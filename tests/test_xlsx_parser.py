"""Phase T6.4.a — multi-sheet xlsx ingestion tests.

Two public functions to cover:

* :func:`core.pricing.xlsx_parser.parse_vendor_xlsx` — bytes in,
  ``dict[sheet_name -> BatchOverridePlan]`` out, one plan per sheet.
* :func:`core.pricing.xlsx_parser.merge_xlsx_plans` — per-sheet plans
  in, one flat plan out with sheet provenance encoded in the row's
  ``notes`` field and ``row_index`` renumbered monotonically.

Plus an end-to-end test: build a 2-sheet workbook with mixed UoMs
(Sheet A = LF, Sheet B = SF), parse → merge → match (T6.4.b
``enforce_uom_compatibility=True``) → apply → assert that
``CostLine.notes`` carries both the ``[vendor-csv]`` source tag (from
T6.4.c) and the ``[sheet: <name>]`` provenance prefix.

Test workbooks are constructed in-memory via ``openpyxl`` so the suite
has zero fixture files on disk.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from core.pricing.batch_override import (
    SOURCE_TAG_MANUAL_OVERRIDE,
    SOURCE_TAG_VENDOR_CSV,
    BatchOverridePlan,
    BatchOverrideRow,
    apply_batch_plan,
    match_cost_lines,
)
from core.pricing.xlsx_parser import (
    merge_xlsx_plans,
    parse_vendor_xlsx,
)
from core.schemas import (
    CostBand,
    CostCategory,
    CostLine,
    CostSourceTier,
    Estimate,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _workbook_bytes(
    sheets: list[tuple[str, list[list[object]]]],
) -> bytes:
    """Build an in-memory xlsx workbook from a list of (sheet_name, rows).

    ``rows`` is a list of cell-value lists — the first list is the
    topmost row, columns left-to-right. Cell values can be any type
    ``openpyxl`` accepts (str, int, float, None).
    """
    wb = Workbook()
    # ``Workbook()`` ships with one default sheet; we drop it so the
    # passed-in list is the source of truth for tab order.
    default_name = wb.active.title
    for sheet_name, rows in sheets:
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(list(row))
    del wb[default_name]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _line(
    *,
    description: str = "Interior latex paint walls",
    csi_division: str = "09",
    csi_section: str = "09 91 23",
    quantity: float = 100.0,
    unit: str = "SF",
    unit_cost: float = 2.0,
) -> CostLine:
    return CostLine(
        csi_division=csi_division,
        csi_section=csi_section,
        description=description,
        quantity=quantity,
        unit=unit,
        unit_cost=unit_cost,
        total_cost=round(unit_cost * quantity, 2),
        cost_category=CostCategory.SUBCONTRACTOR,
        confidence=0.9,
        price_confidence=0.7,
        cost_source_tier=CostSourceTier.INTERPOLATED,
        cost_band=CostBand.OPERATOR_REVIEW,
        suppressed=False,
        cost_source="cwicr:test",
        notes=None,
    )


def _estimate(lines: list[CostLine]) -> Estimate:
    return Estimate(
        project_name="T6.4.a xlsx ingestion",
        region_multiplier=1.0,
        contingency_pct=10.0,
        overhead_pct=10.0,
        profit_pct=5.0,
        line_items=lines,
    )


def _plan_rows(plan: BatchOverridePlan) -> list[BatchOverrideRow]:
    """Extract rows from any of a plan's buckets (matcher hasn't run)."""
    out = (
        list(plan.matched)
        + list(plan.ambiguous)
        + list(plan.low_similarity)
        + list(plan.no_match)
    )
    out.sort(key=lambda r: r.row.row_index)
    return [r.row for r in out]


# ---------------------------------------------------------------------------
# parse_vendor_xlsx — single sheet / multi sheet / order
# ---------------------------------------------------------------------------


class TestParseVendorXlsxBasic:
    def test_single_sheet_with_valid_header_parses_one_plan(self) -> None:
        wb = _workbook_bytes([
            ("Pricing", [
                ["description", "unit_cost"],
                ["Interior paint", 2.50],
                ["Slab on grade", 8.00],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        assert list(plans.keys()) == ["Pricing"]
        plan = plans["Pricing"]
        assert plan.total_rows == 2
        rows = _plan_rows(plan)
        assert rows[0].description == "Interior paint"
        assert rows[0].unit_cost == pytest.approx(2.50)
        assert rows[1].description == "Slab on grade"

    def test_multi_sheet_parses_three_plans_in_insertion_order(self) -> None:
        wb = _workbook_bytes([
            ("Mech", [
                ["description", "unit_cost"],
                ["AHU-1", 5000.0],
            ]),
            ("Elec", [
                ["description", "unit_cost"],
                ["Panelboard 200A", 1200.0],
                ["Wire pull 12 AWG", 0.85],
            ]),
            ("Plumb", [
                ["description", "unit_cost"],
                ["Lavatory P-1", 450.0],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        assert list(plans.keys()) == ["Mech", "Elec", "Plumb"]
        assert plans["Mech"].total_rows == 1
        assert plans["Elec"].total_rows == 2
        assert plans["Plumb"].total_rows == 1

    def test_dict_returned_preserves_workbook_sheet_order(self) -> None:
        wb = _workbook_bytes([
            ("Z-Last", [["description", "unit_cost"], ["x", 1.0]]),
            ("A-First", [["description", "unit_cost"], ["y", 2.0]]),
            ("M-Middle", [["description", "unit_cost"], ["z", 3.0]]),
        ])
        plans = parse_vendor_xlsx(wb)
        # Insertion order is workbook tab order, NOT alphabetical.
        assert list(plans.keys()) == ["Z-Last", "A-First", "M-Middle"]


# ---------------------------------------------------------------------------
# parse_vendor_xlsx — sheet skipping rules
# ---------------------------------------------------------------------------


class TestParseVendorXlsxSheetSkipping:
    def test_empty_sheet_skipped(self) -> None:
        wb = _workbook_bytes([
            ("Empty", []),
            ("Real", [
                ["description", "unit_cost"],
                ["Interior paint", 2.50],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        assert list(plans.keys()) == ["Real"]

    def test_sheet_with_only_one_column_skipped(self) -> None:
        wb = _workbook_bytes([
            ("SingleCol", [["description"], ["solo"], ["foo"]]),
            ("Real", [
                ["description", "unit_cost"],
                ["Interior paint", 2.50],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        assert "SingleCol" not in plans
        assert "Real" in plans

    def test_sheet_with_no_recognised_columns_skipped(self) -> None:
        wb = _workbook_bytes([
            ("CoverSheet", [
                ["Acme Corp Quote", "2026-05-28"],
                ["For: Client Co.", "Job #99"],
            ]),
            ("Real", [
                ["description", "unit_cost"],
                ["Interior paint", 2.50],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        assert "CoverSheet" not in plans
        assert plans["Real"].total_rows == 1

    def test_banner_row_above_header_detected(self) -> None:
        wb = _workbook_bytes([
            ("WithBanner", [
                ["ACME Plumbing", None, None],         # banner row 1
                ["Quote #QT-2026-05-28", None, None],   # banner row 2
                ["description", "unit_cost", "vendor"],
                ["Lavatory P-1", 450.0, "ACME"],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        plan = plans["WithBanner"]
        rows = _plan_rows(plan)
        assert len(rows) == 1
        assert rows[0].description == "Lavatory P-1"
        assert rows[0].vendor == "ACME"

    def test_sheet_with_header_but_no_data_kept_as_empty_plan(self) -> None:
        wb = _workbook_bytes([
            ("HeaderOnly", [
                ["description", "unit_cost"],
            ]),
            ("Real", [
                ["description", "unit_cost"],
                ["Interior paint", 2.50],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        assert "HeaderOnly" in plans
        assert plans["HeaderOnly"].total_rows == 0
        assert plans["Real"].total_rows == 1

    def test_all_sheets_empty_raises(self) -> None:
        wb = _workbook_bytes([
            ("E1", []),
            ("E2", []),
        ])
        with pytest.raises(ValueError, match="zero usable sheets"):
            parse_vendor_xlsx(wb)


# ---------------------------------------------------------------------------
# parse_vendor_xlsx — column alias / uom interop
# ---------------------------------------------------------------------------


class TestParseVendorXlsxColumns:
    def test_column_aliases_desc_price_qty(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["desc", "price", "qty"],
                ["Interior paint", 2.50, 100],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        rows = _plan_rows(plans["Sheet1"])
        assert rows[0].description == "Interior paint"
        assert rows[0].unit_cost == pytest.approx(2.50)
        assert rows[0].quantity == pytest.approx(100.0)

    def test_case_insensitive_headers(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["DESCRIPTION", "UNIT_COST"],
                ["Interior paint", 2.50],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        assert plans["Sheet1"].total_rows == 1

    def test_uom_column_picked_up_for_t64b_interop(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["description", "unit_cost", "uom"],
                ["Fence post LF priced", 12.0, "LF"],
                ["Roof TPO SF priced", 9.50, "SF"],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        rows = _plan_rows(plans["Sheet1"])
        assert rows[0].unit_of_measure == "LF"
        assert rows[1].unit_of_measure == "SF"

    def test_unit_of_measure_alias_picked_up(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["description", "unit_cost", "unit"],
                ["Fence post", 12.0, "linear ft"],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        rows = _plan_rows(plans["Sheet1"])
        assert rows[0].unit_of_measure == "linear ft"

    def test_extended_derives_unit_cost(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["description", "extended", "qty"],
                ["Misc framing", 1250.00, 100],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        rows = _plan_rows(plans["Sheet1"])
        assert rows[0].unit_cost == pytest.approx(12.50)


# ---------------------------------------------------------------------------
# parse_vendor_xlsx — per-row error handling
# ---------------------------------------------------------------------------


class TestParseVendorXlsxRowErrors:
    def test_negative_unit_cost_skipped(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["description", "unit_cost"],
                ["Bad row", -5.0],
                ["Good row", 8.00],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        rows = _plan_rows(plans["Sheet1"])
        assert len(rows) == 1
        assert rows[0].description == "Good row"

    def test_empty_description_skipped(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["description", "unit_cost"],
                ["", 2.50],
                ["Good row", 8.00],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        rows = _plan_rows(plans["Sheet1"])
        assert len(rows) == 1
        assert rows[0].description == "Good row"

    def test_dollar_signs_and_commas_in_cells(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [
                ["description", "unit_cost"],
                ["Big ticket", "$1,250.00"],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        rows = _plan_rows(plans["Sheet1"])
        assert rows[0].unit_cost == pytest.approx(1250.00)


# ---------------------------------------------------------------------------
# parse_vendor_xlsx — sheet_filter
# ---------------------------------------------------------------------------


class TestParseVendorXlsxSheetFilter:
    def test_sheet_filter_callable_applied(self) -> None:
        wb = _workbook_bytes([
            ("Cover", [["description", "unit_cost"], ["x", 1.0]]),
            ("Mech", [["description", "unit_cost"], ["y", 2.0]]),
            ("Summary", [["description", "unit_cost"], ["z", 3.0]]),
        ])
        plans = parse_vendor_xlsx(
            wb, sheet_filter=lambda name: name == "Mech"
        )
        assert list(plans.keys()) == ["Mech"]

    def test_sheet_filter_returns_false_for_all_raises(self) -> None:
        wb = _workbook_bytes([
            ("Sheet1", [["description", "unit_cost"], ["x", 1.0]]),
        ])
        with pytest.raises(ValueError, match="zero usable sheets"):
            parse_vendor_xlsx(wb, sheet_filter=lambda name: False)

    def test_sheet_filter_raising_treats_sheet_as_filtered(self) -> None:
        def bad_filter(name: str) -> bool:
            if name == "Bad":
                raise RuntimeError("filter exploded")
            return True

        wb = _workbook_bytes([
            ("Bad", [["description", "unit_cost"], ["a", 1.0]]),
            ("Good", [["description", "unit_cost"], ["b", 2.0]]),
        ])
        plans = parse_vendor_xlsx(wb, sheet_filter=bad_filter)
        assert list(plans.keys()) == ["Good"]


# ---------------------------------------------------------------------------
# parse_vendor_xlsx — invalid input
# ---------------------------------------------------------------------------


class TestParseVendorXlsxInvalidInput:
    def test_empty_bytes_raises(self) -> None:
        with pytest.raises(ValueError, match="empty"):
            parse_vendor_xlsx(b"")

    def test_garbage_bytes_raises(self) -> None:
        with pytest.raises(ValueError):
            parse_vendor_xlsx(b"not an xlsx file at all")


# ---------------------------------------------------------------------------
# parse_vendor_xlsx — per_sheet_tag_override parameter accepted
# ---------------------------------------------------------------------------


class TestParseVendorXlsxPerSheetTag:
    def test_per_sheet_tag_override_accepted_without_error(self) -> None:
        """The kwarg routes through the UI; the parser accepts and ignores it.

        The override is consumed by the Streamlit UI when it calls
        :func:`apply_batch_plan` (which takes the ``source_tag`` kwarg).
        At the parser level we only assert the kwarg is accepted and
        doesn't change parse output.
        """
        wb = _workbook_bytes([
            ("Sheet1", [["description", "unit_cost"], ["x", 1.0]]),
        ])
        plans = parse_vendor_xlsx(
            wb,
            per_sheet_tag_override={"Sheet1": SOURCE_TAG_MANUAL_OVERRIDE},
        )
        assert plans["Sheet1"].total_rows == 1


# ---------------------------------------------------------------------------
# merge_xlsx_plans — row count / renumbering / provenance
# ---------------------------------------------------------------------------


class TestMergeXlsxPlans:
    def _three_sheet_workbook(self) -> bytes:
        return _workbook_bytes([
            ("A", [
                ["description", "unit_cost"],
                ["Row A1", 1.0],
                ["Row A2", 2.0],
            ]),
            ("B", [
                ["description", "unit_cost"],
                ["Row B1", 10.0],
            ]),
            ("C", [
                ["description", "unit_cost"],
                ["Row C1", 100.0],
                ["Row C2", 200.0],
                ["Row C3", 300.0],
            ]),
        ])

    def test_three_plans_flatten_to_n1_plus_n2_plus_n3_rows(self) -> None:
        plans = parse_vendor_xlsx(self._three_sheet_workbook())
        merged = merge_xlsx_plans(plans)
        assert merged.total_rows == 2 + 1 + 3

    def test_merged_row_indices_unique_and_monotonic(self) -> None:
        plans = parse_vendor_xlsx(self._three_sheet_workbook())
        merged = merge_xlsx_plans(plans)
        indices = [r.row.row_index for r in merged.no_match]
        assert indices == sorted(indices)
        assert len(set(indices)) == len(indices)
        assert indices[0] == 2  # mirrors CSV header = row 1

    def test_merged_csv_row_identifier_prefixed_with_sheet_name(self) -> None:
        plans = parse_vendor_xlsx(self._three_sheet_workbook())
        merged = merge_xlsx_plans(plans, sheet_separator=" :: ")
        # Each row's notes carries the sheet name + separator at the
        # front, so the merged plan can be audited back to the source
        # sheet even though row_index is renumbered.
        first_a = merged.no_match[0].row
        assert first_a.notes is not None
        assert first_a.notes.startswith("A :: ")
        first_c = merged.no_match[3].row
        assert first_c.notes is not None
        assert first_c.notes.startswith("C :: ")

    def test_custom_separator_honoured(self) -> None:
        plans = parse_vendor_xlsx(self._three_sheet_workbook())
        merged = merge_xlsx_plans(plans, sheet_separator=" | ")
        first = merged.no_match[0].row
        assert first.notes is not None
        assert first.notes.startswith("A | ")

    def test_existing_notes_preserved_after_provenance_prefix(self) -> None:
        wb = _workbook_bytes([
            ("S1", [
                ["description", "unit_cost", "notes"],
                ["row", 1.0, "FOB origin"],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        merged = merge_xlsx_plans(plans)
        notes = merged.no_match[0].row.notes
        assert notes is not None
        assert notes.startswith("S1 :: ")
        assert "FOB origin" in notes

    def test_empty_input_dict_raises(self) -> None:
        with pytest.raises(ValueError, match="zero plans"):
            merge_xlsx_plans({})

    def test_single_sheet_input_no_double_prefix(self) -> None:
        wb = _workbook_bytes([
            ("Solo", [
                ["description", "unit_cost"],
                ["only row", 1.0],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        merged = merge_xlsx_plans(plans)
        notes = merged.no_match[0].row.notes
        assert notes is not None
        assert notes.count("Solo :: ") == 1

    def test_merged_descriptions_and_unit_costs_preserved(self) -> None:
        plans = parse_vendor_xlsx(self._three_sheet_workbook())
        merged = merge_xlsx_plans(plans)
        descriptions = [r.row.description for r in merged.no_match]
        assert descriptions == [
            "Row A1", "Row A2",
            "Row B1",
            "Row C1", "Row C2", "Row C3",
        ]
        costs = [r.row.unit_cost for r in merged.no_match]
        assert costs == [1.0, 2.0, 10.0, 100.0, 200.0, 300.0]


# ---------------------------------------------------------------------------
# End-to-end — parse → merge → match → apply with UoM safety + provenance
# ---------------------------------------------------------------------------


class TestEndToEndXlsxToCostLineNotes:
    def test_two_sheet_mixed_uom_no_cross_application(self) -> None:
        """Sheet A = LF prices, Sheet B = SF prices. T6.4.b's
        ``enforce_uom_compatibility=True`` (default) must route each
        row only to its UoM-compatible cost line; no LF row should
        land on an SF line and vice versa.
        """
        wb = _workbook_bytes([
            ("LinearFeet", [
                ["description", "unit_cost", "uom"],
                ["Fence post galvanised", 12.0, "LF"],
            ]),
            ("SquareFeet", [
                ["description", "unit_cost", "uom"],
                ["Roof membrane TPO", 9.50, "SF"],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        merged = merge_xlsx_plans(plans)

        # Build a cost-line set with both an LF line and an SF line so
        # the matcher has both UoMs to choose from.
        lines = [
            _line(
                description="Fence post galvanised",
                unit="LF",
                csi_section="32 31 13",
                unit_cost=8.0,
            ),
            _line(
                description="Roof membrane TPO",
                unit="SF",
                csi_section="07 54 23",
                unit_cost=7.0,
            ),
        ]
        rows = [r.row for r in merged.no_match]
        match_plan = match_cost_lines(
            rows, lines, enforce_uom_compatibility=True
        )

        assert len(match_plan.matched) == 2
        match_by_desc = {
            r.row.description: r for r in match_plan.matched
        }
        assert match_by_desc["Fence post galvanised"].best_match_index == 0
        assert match_by_desc["Roof membrane TPO"].best_match_index == 1

    def test_apply_propagates_vendor_csv_tag_and_sheet_provenance(self) -> None:
        wb = _workbook_bytes([
            ("Pricing", [
                ["description", "unit_cost"],
                ["Interior latex paint walls", 3.50],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        merged = merge_xlsx_plans(plans)
        lines = [_line(description="Interior latex paint walls", unit_cost=2.0)]
        rows = [r.row for r in merged.no_match]
        match_plan = match_cost_lines(rows, lines)
        assert len(match_plan.matched) == 1
        est = _estimate(lines)
        new_est, summary = apply_batch_plan(
            est, match_plan, source_tag=SOURCE_TAG_VENDOR_CSV
        )
        assert new_est.line_items[0].unit_cost == pytest.approx(3.50)
        notes = new_est.line_items[0].notes or ""
        # T6.4.c tag at position 0.
        assert notes.startswith(SOURCE_TAG_VENDOR_CSV + " ")
        # T6.4.a sheet provenance preserved somewhere in the notes
        # (the merge step prepended "Pricing :: row 2" to the row's
        # notes; that string in turn lands in the operator note via
        # ``format_batch_operator_note``).
        assert "Pricing :: row 2" in notes

    def test_no_match_rows_remain_in_no_match_after_merge(self) -> None:
        """Rows whose descriptions don't match any cost line should
        still cleanly bucket as NO_MATCH after the merge step. The
        merge function is structural — it preserves row identity, not
        match status."""
        wb = _workbook_bytes([
            ("Sheet1", [
                ["description", "unit_cost"],
                ["zzz nonsense xxxx", 99.0],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        merged = merge_xlsx_plans(plans)
        lines = [_line(description="Interior latex paint walls", unit_cost=2.0)]
        rows = [r.row for r in merged.no_match]
        match_plan = match_cost_lines(rows, lines)
        assert len(match_plan.no_match) == 1
        # NOT applied even when apply runs.
        est = _estimate(lines)
        new_est, _ = apply_batch_plan(est, match_plan)
        assert new_est.line_items[0].unit_cost == pytest.approx(2.0)


# ---------------------------------------------------------------------------
# Per-sheet UoM mismatch end-to-end — UoM gate works at merge time
# ---------------------------------------------------------------------------


class TestMixedUomMatcherInterop:
    def test_lf_row_does_not_match_sf_only_lines(self) -> None:
        """A merged row carrying LF UoM must NOT match against a cost
        line in SF when ``enforce_uom_compatibility=True``. Verifies
        that the UoM column from the xlsx survives merge and reaches
        the matcher."""
        wb = _workbook_bytes([
            ("LF_only", [
                ["description", "unit_cost", "uom"],
                ["Custom widget", 12.0, "LF"],
            ]),
        ])
        plans = parse_vendor_xlsx(wb)
        merged = merge_xlsx_plans(plans)
        rows = [r.row for r in merged.no_match]
        # Cost line with the same description but SF — should NOT match
        # under safety-on.
        lines = [
            _line(
                description="Custom widget",
                unit="SF",
                unit_cost=5.0,
            ),
        ]
        match_plan = match_cost_lines(
            rows, lines, enforce_uom_compatibility=True
        )
        assert len(match_plan.matched) == 0
        assert len(match_plan.no_match) == 1
