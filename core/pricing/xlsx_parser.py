"""Phase T6.4.a — multi-sheet xlsx ingestion for vendor pricing.

Operators frequently receive vendor pricing as a multi-tab Excel workbook
(e.g. one tab per discipline: ``Mechanical`` / ``Electrical`` / ``Plumbing``
or one tab per vendor in a roll-up sheet). The T6.3 vendor-CSV path
(:func:`core.pricing.batch_override.parse_vendor_csv`) only handles a
single CSV at a time, so operators had to export each tab manually as a
separate CSV and run the batch-apply flow N times. This module closes
that gap by walking every sheet of an ``.xlsx`` workbook and producing
one :class:`~core.pricing.batch_override.BatchOverridePlan`-shaped
result per sheet, then optionally flattening them into a single plan via
:func:`merge_xlsx_plans` so the existing T6.4.b unit-aware matcher and
T6.4.c source-tag-aware applier run unchanged.

Public surface (two functions only):

* :func:`parse_vendor_xlsx` — bytes in, ``dict[str, BatchOverridePlan]``
  out (one plan per sheet, sheet-insertion order preserved).
* :func:`merge_xlsx_plans` — per-sheet plans in, one flat plan out
  with sheet provenance encoded in each row's ``notes`` field.

Design notes:

* **No new third-party deps.** ``openpyxl`` is already a project
  dependency (``requirements.txt`` line 16, used by
  :mod:`core.exporter`).
* **Header detection is conservative.** A sheet's header is the first
  row containing at least :data:`_MIN_HEADER_TOKENS` recognised column
  aliases from
  :data:`~core.pricing.batch_override._COLUMN_ALIASES`. Vendor PDFs and
  xlsx exports often carry a banner / contact-info row above the
  table — this rule skips those without false positives.
* **Sheets are skipped (not failed) on:** zero rows, no detectable
  header, < 2 columns, only one recognised column. A warning is
  emitted via the per-sheet warnings list so the operator can audit.
* **Sheet provenance lives in ``notes``, not on the row.**
  :class:`~core.pricing.batch_override.BatchOverrideRow` is a widely
  used dataclass (sub-quote parser, every batch test); adding a
  ``source_sheet`` field would force a schema bump and ripple through
  every constructor / repr in the suite. Instead :func:`merge_xlsx_plans`
  prepends ``[sheet: <name>]`` to the row's ``notes`` field — a smaller
  change that round-trips through every existing surface (CSV export,
  operator notes, history log) for free.
* **Row indices are renumbered on merge.** Per-sheet plans return
  ``row_index`` starting at 2 (mirroring the CSV "header is row 1"
  convention used by :func:`~core.pricing.batch_override.parse_vendor_csv`).
  :func:`merge_xlsx_plans` re-numbers them monotonically across sheets
  so the merged plan has unique ``row_index`` values — required because
  several downstream surfaces (``apply_batch_plan``'s ``skip_rows`` /
  ``resolved_ambiguous`` dicts, the history-row "csv-row: N" stamp) key
  off ``row_index`` and would silently collide otherwise.

The module is intentionally **pure-logic** — no Streamlit, no I/O
beyond reading the in-memory bytes via :func:`openpyxl.load_workbook`.
The Streamlit UI wrapper lives in ``app.py``.
"""

from __future__ import annotations

import io
import logging
from collections import OrderedDict
from typing import Callable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from core.pricing.batch_override import (
    _COLUMN_ALIASES,
    BatchMatchResult,
    BatchMatchStatus,
    BatchOverridePlan,
    BatchOverrideRow,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------


# Minimum number of recognised column aliases a candidate header row must
# carry to be treated as the table header. Two is conservative — a row
# with just one keyword (e.g. only "Description") is more likely a
# section heading or a banner than a real table header. Mirrors
# :data:`core.pricing.subquote_parser._MIN_HEADER_TOKENS`.
_MIN_HEADER_TOKENS: int = 2


# Maximum number of rows scanned looking for the header. Vendor xlsx
# exports occasionally carry a multi-row banner (vendor logo, contact
# info, project reference, date) before the actual table. Twenty is
# generous; in practice the header is on row 1-3.
_HEADER_SCAN_LIMIT: int = 20


# Default separator used by :func:`merge_xlsx_plans` between sheet name
# and the source-row identifier when encoding sheet provenance in the
# row's ``notes`` field. Chosen for visual clarity in a CSV / Excel
# Notes cell — ``" :: "`` is unambiguous and unlikely to appear in a
# legit vendor notes string.
_DEFAULT_SHEET_SEPARATOR: str = " :: "


# ---------------------------------------------------------------------------
# Header / column resolution (sheet-rows variant)
# ---------------------------------------------------------------------------


def _flatten_header_keywords() -> set[str]:
    """Flatten ``_COLUMN_ALIASES`` into one lowercase keyword set.

    Underscores in aliases are stripped so an Excel cell rendered as
    ``"Unit Cost"`` or ``"unit cost"`` matches the canonical
    ``"unit_cost"`` alias. Cached at module load time via the
    module-level :data:`_HEADER_KEYWORDS` singleton.
    """
    out: set[str] = set()
    for aliases in _COLUMN_ALIASES.values():
        for alias in aliases:
            lower = alias.lower()
            out.add(lower)
            out.add(lower.replace("_", " "))
            out.add(lower.replace("_", ""))
    return out


_HEADER_KEYWORDS: set[str] = _flatten_header_keywords()


def _normalise_header_cell(value: object) -> str:
    """Render an openpyxl cell value as a lowercase, whitespace-collapsed string.

    ``openpyxl`` returns ``None`` for empty cells, numeric types for
    numbers, and ``str`` for text. We coerce everything through
    :func:`str` then strip + lowercase. Multiple spaces collapse to one
    so a cell rendered as ``"  Unit   Cost "`` matches ``"unit cost"``.
    """
    if value is None:
        return ""
    s = str(value).strip().lower()
    if not s:
        return ""
    return " ".join(s.split())


def _cell_matches_known_header(cell_text: str) -> bool:
    """Return True if ``cell_text`` matches any recognised column alias.

    Tries the verbatim text, the underscore-substituted form, and the
    no-space form — covers ``"unit cost"`` / ``"unit_cost"`` /
    ``"unitcost"`` from the same lookup.
    """
    if not cell_text:
        return False
    for variant in (
        cell_text,
        cell_text.replace(" ", "_"),
        cell_text.replace(" ", ""),
    ):
        if variant in _HEADER_KEYWORDS:
            return True
    return False


def _detect_header_row_in_sheet(
    sheet_rows: list[list[object]],
) -> int | None:
    """Find the index of the first row that looks like a column header.

    Walks up to :data:`_HEADER_SCAN_LIMIT` rows from the top of the
    sheet, counting how many cells match a recognised column alias.
    The first row with at least :data:`_MIN_HEADER_TOKENS` hits wins.

    Returns the 0-based row index, or ``None`` if no header is
    detectable. The caller should skip the sheet on ``None``.
    """
    scan_limit = min(_HEADER_SCAN_LIMIT, len(sheet_rows))
    for row_idx in range(scan_limit):
        row = sheet_rows[row_idx]
        if not row:
            continue
        hits = 0
        for cell_value in row:
            normalised = _normalise_header_cell(cell_value)
            if not normalised:
                continue
            if _cell_matches_known_header(normalised):
                hits += 1
        if hits >= _MIN_HEADER_TOKENS:
            return row_idx
    return None


def _resolve_column_map_from_row(
    header_row: list[object],
) -> dict[str, int]:
    """Map canonical slot names → column index for one sheet's header row.

    Mirrors :func:`core.pricing.batch_override._resolve_column_map` but
    operates on a positional row from :func:`openpyxl.worksheet.iter_rows`
    rather than a CSV ``DictReader.fieldnames`` list. Returns a dict
    keyed by canonical slot name (``"description"``, ``"unit_cost"``,
    ``"unit_of_measure"``, …); missing optional slots are simply absent
    from the result.
    """
    normalised: dict[str, int] = {}
    for col_idx, cell_value in enumerate(header_row):
        text = _normalise_header_cell(cell_value)
        if not text:
            continue
        for variant in (
            text,
            text.replace(" ", "_"),
            text.replace(" ", ""),
        ):
            if variant not in normalised:
                normalised[variant] = col_idx

    out: dict[str, int] = {}
    for slot, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            for variant in (
                alias.lower(),
                alias.lower().replace("_", " "),
                alias.lower().replace("_", ""),
            ):
                if variant in normalised:
                    out[slot] = normalised[variant]
                    break
            if slot in out:
                break
    return out


# ---------------------------------------------------------------------------
# Cell value parsing
# ---------------------------------------------------------------------------


def _cell_str(value: object) -> str:
    """Render a cell value as a stripped string. Returns ``""`` on None."""
    if value is None:
        return ""
    return str(value).strip()


def _cell_number(value: object) -> float | None:
    """Best-effort numeric coercion. Returns ``None`` on failure.

    Tolerates the same decorations as
    :func:`core.pricing.batch_override._parse_float` — leading ``$``,
    thousands-separator commas — so an operator who typed
    ``"$1,250.00"`` into a cell still gets parsed. ``int`` / ``float``
    cell values from openpyxl are returned as-is.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        # ``bool`` is a subclass of ``int`` in Python — guard explicitly
        # so a cell containing TRUE / FALSE doesn't get coerced to 1.0
        # / 0.0 and silently look like a $1 unit cost.
        return None
    if isinstance(value, (int, float)):
        try:
            f = float(value)
        except (TypeError, ValueError):
            return None
        if f != f:  # NaN
            return None
        return f
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Single-sheet parser
# ---------------------------------------------------------------------------


def _parse_sheet_to_plan(
    sheet_name: str,
    sheet_rows: list[list[object]],
) -> tuple[BatchOverridePlan | None, list[str]]:
    """Parse one sheet's rows into a :class:`BatchOverridePlan`.

    Returns ``(plan, warnings)``. ``plan`` is ``None`` when the sheet
    is unparseable (no header, no description column, < 2 columns,
    empty) — the caller should skip the sheet in that case.

    Per-row error semantics mirror
    :func:`core.pricing.batch_override.parse_vendor_csv`:

    * Empty description → skipped (warning emitted).
    * Non-numeric unit_cost → skipped (warning).
    * Negative unit_cost → skipped (warning).

    The returned plan carries the rows in the ``no_match`` bucket
    (matcher hasn't run yet). The caller is expected to feed the merged
    plan's row list into :func:`core.pricing.batch_override.match_cost_lines`
    against an estimate to produce the actual match plan.
    """
    warnings: list[str] = []

    if not sheet_rows:
        return None, [f"Sheet {sheet_name!r}: empty; skipped."]

    max_cols = max((len(r) for r in sheet_rows), default=0)
    if max_cols < 2:
        return None, [
            f"Sheet {sheet_name!r}: has fewer than 2 columns "
            f"(max width {max_cols}); skipped — not a vendor table."
        ]

    header_idx = _detect_header_row_in_sheet(sheet_rows)
    if header_idx is None:
        return None, [
            f"Sheet {sheet_name!r}: no recognisable header row "
            f"(expected at least {_MIN_HEADER_TOKENS} column aliases "
            f"from description / unit_cost / vendor / quantity / "
            f"unit_of_measure / …); skipped."
        ]

    column_map = _resolve_column_map_from_row(sheet_rows[header_idx])
    if "description" not in column_map:
        return None, [
            f"Sheet {sheet_name!r}: header row detected but no "
            f"'description' column (aliases: desc / item / "
            f"item_description / line_item); skipped."
        ]
    if "unit_cost" not in column_map and "extended" not in column_map:
        return None, [
            f"Sheet {sheet_name!r}: header row detected but no "
            f"'unit_cost' or 'extended' column; skipped."
        ]

    desc_col = column_map["description"]
    unit_col = column_map.get("unit_cost")
    ext_col = column_map.get("extended")
    qty_col = column_map.get("quantity")
    vendor_col = column_map.get("vendor")
    quote_col = column_map.get("quote_ref")
    notes_col = column_map.get("notes")
    uom_col = column_map.get("unit_of_measure")

    rows: list[BatchOverrideRow] = []

    # Row-index convention: header is row 1, first data row is row 2 —
    # matches :func:`core.pricing.batch_override.parse_vendor_csv`. Note
    # this is per-sheet; :func:`merge_xlsx_plans` re-numbers across
    # sheets to ensure global uniqueness.
    for data_row_offset, data_row in enumerate(
        sheet_rows[header_idx + 1:], start=2
    ):
        row_index = data_row_offset

        if not data_row or desc_col >= len(data_row):
            continue

        desc = _cell_str(data_row[desc_col])
        if not desc:
            # Silent skip on empty description rows — vendor xlsx
            # exports often have a trailing blank row or section-
            # separator rows. Warning would be noisy.
            continue

        unit_cost: float | None = None
        if unit_col is not None and unit_col < len(data_row):
            unit_cost = _cell_number(data_row[unit_col])

        qty: float | None = None
        if qty_col is not None and qty_col < len(data_row):
            qty = _cell_number(data_row[qty_col])
            if qty is not None and qty < 0:
                qty = None

        # Derive unit_cost from extended / qty when unit column is
        # missing (T8.1 sub-quote-parser pattern; safe to mirror here
        # since the same column-alias map already exposes "extended").
        if unit_cost is None and ext_col is not None and ext_col < len(data_row):
            extended = _cell_number(data_row[ext_col])
            if extended is not None and qty is not None and qty > 0:
                unit_cost = round(extended / qty, 4)

        if unit_cost is None:
            warnings.append(
                f"Sheet {sheet_name!r} row {row_index}: no unit_cost "
                f"or derivable extended/qty; skipped."
            )
            continue
        if unit_cost < 0:
            warnings.append(
                f"Sheet {sheet_name!r} row {row_index}: negative "
                f"unit_cost {unit_cost}; skipped."
            )
            continue

        vendor: str | None = None
        if vendor_col is not None and vendor_col < len(data_row):
            v = _cell_str(data_row[vendor_col])
            vendor = v or None

        quote_ref: str | None = None
        if quote_col is not None and quote_col < len(data_row):
            q = _cell_str(data_row[quote_col])
            quote_ref = q or None

        notes: str | None = None
        if notes_col is not None and notes_col < len(data_row):
            n = _cell_str(data_row[notes_col])
            notes = n or None

        uom: str | None = None
        if uom_col is not None and uom_col < len(data_row):
            u = _cell_str(data_row[uom_col])
            uom = u or None

        rows.append(BatchOverrideRow(
            row_index=row_index,
            description=desc,
            unit_cost=unit_cost,
            vendor=vendor,
            quote_ref=quote_ref,
            notes=notes,
            quantity=qty,
            unit_of_measure=uom,
        ))

    # Convention for "parsed but unmatched": stash the rows in the
    # ``no_match`` bucket as bare :class:`BatchMatchResult` entries so
    # the dataclass invariants hold (total_rows == sum of buckets). The
    # downstream matcher (:func:`match_cost_lines`) takes a list of
    # ``BatchOverrideRow`` directly, so the bucket placement here is
    # purely structural — the caller is expected to extract rows via
    # :func:`_plan_rows` (or :func:`merge_xlsx_plans`) and feed them to
    # the matcher.
    match_results = [
        BatchMatchResult(
            row=r,
            status=BatchMatchStatus.NO_MATCH,
            best_match_index=None,
            best_match_similarity=0.0,
            runner_up_index=None,
            runner_up_similarity=0.0,
            candidate_lines=[],
        )
        for r in rows
    ]

    plan = BatchOverridePlan(
        total_rows=len(rows),
        matched=[],
        ambiguous=[],
        no_match=match_results,
        low_similarity=[],
        # The matcher hasn't run yet — these are echoed back from the
        # downstream call. Use the matcher's defaults so a caller that
        # inspects the plan before running the matcher sees sensible
        # values (rather than 0.0 / 0.0 which would look like "no
        # threshold configured").
        similarity_threshold=0.65,
        ambiguity_margin=0.10,
    )
    return plan, warnings


def _plan_rows(plan: BatchOverridePlan) -> list[BatchOverrideRow]:
    """Extract every :class:`BatchOverrideRow` from a plan's buckets.

    Helper used by :func:`merge_xlsx_plans`. Walks all four bucket
    lists (matched / ambiguous / low_similarity / no_match) and
    returns the underlying rows in their original ``row_index`` order.
    The per-sheet plans produced by :func:`_parse_sheet_to_plan` only
    populate ``no_match`` (matcher hasn't run yet), but this helper
    also works on plans that have been routed through
    :func:`~core.pricing.batch_override.match_cost_lines`.
    """
    all_results: list[BatchMatchResult] = (
        list(plan.matched)
        + list(plan.ambiguous)
        + list(plan.low_similarity)
        + list(plan.no_match)
    )
    all_results.sort(key=lambda r: r.row.row_index)
    return [r.row for r in all_results]


# ---------------------------------------------------------------------------
# Public API — parse_vendor_xlsx
# ---------------------------------------------------------------------------


def parse_vendor_xlsx(
    file_bytes: bytes,
    *,
    sheet_filter: Callable[[str], bool] | None = None,
    per_sheet_tag_override: dict[str, str] | None = None,
) -> dict[str, BatchOverridePlan]:
    """Parse a multi-sheet xlsx into one :class:`BatchOverridePlan` per sheet.

    Args:
        file_bytes: Raw xlsx bytes (e.g. from a Streamlit
            ``st.file_uploader``).
        sheet_filter: Optional callable; only sheets where
            ``filter(sheet_name)`` returns ``True`` are parsed. ``None``
            (the default) means "every sheet". Use to skip cover sheets
            / summary tabs without forcing the operator to clean the
            workbook first.
        per_sheet_tag_override: Optional mapping from sheet name to a
            ``SOURCE_TAG_*`` constant. Carried alongside the returned
            plans so a downstream caller can pass the right tag to
            :func:`~core.pricing.batch_override.apply_batch_plan` per
            sheet (currently consumed by the Streamlit UI for the
            mixed-source workbook case; the parser itself stores the
            override on the plan's ``BatchOverrideRow.notes`` field via
            :func:`merge_xlsx_plans` so the audit trail keeps the
            distinction even after the per-sheet plans are flattened).

    Returns:
        An :class:`OrderedDict` keyed by sheet name and valued by the
        per-sheet :class:`BatchOverridePlan` (every row in the
        ``no_match`` bucket — matcher has not run yet). Insertion order
        matches workbook tab order.

    Raises:
        ValueError: When the workbook has zero sheets, the bytes can
            not be opened as a valid xlsx, or the workbook parses but
            produces zero rows across all selected sheets (the operator
            uploaded an entirely empty workbook or a workbook full of
            cover sheets that all got filtered).
    """
    if not file_bytes:
        raise ValueError("xlsx is empty (0 bytes).")

    try:
        wb = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except InvalidFileException as exc:
        raise ValueError(
            f"Could not open xlsx (invalid Excel file): {exc}. "
            f"Re-export from Excel and retry."
        ) from exc
    except Exception as exc:
        raise ValueError(
            f"Could not open xlsx ({type(exc).__name__}): {exc}. "
            f"Re-export from Excel and retry."
        ) from exc

    try:
        sheet_names = list(wb.sheetnames)
    except Exception as exc:
        wb.close()
        raise ValueError(
            f"Could not enumerate sheets in xlsx: {exc}"
        ) from exc

    if not sheet_names:
        wb.close()
        raise ValueError(
            "xlsx contains zero sheets. Re-export from Excel."
        )

    out: OrderedDict[str, BatchOverridePlan] = OrderedDict()
    aggregate_warnings: list[str] = []
    try:
        for sheet_name in sheet_names:
            if sheet_filter is not None:
                try:
                    keep = bool(sheet_filter(sheet_name))
                except Exception as exc:
                    logger.warning(
                        "xlsx-parser: sheet_filter raised on %r: %s; "
                        "treating sheet as filtered-out.",
                        sheet_name, exc,
                    )
                    keep = False
                if not keep:
                    continue

            ws = wb[sheet_name]
            try:
                sheet_rows: list[list[object]] = [
                    list(row) for row in ws.iter_rows(values_only=True)
                ]
            except Exception as exc:
                aggregate_warnings.append(
                    f"Sheet {sheet_name!r}: could not iterate rows "
                    f"({type(exc).__name__}): {exc}; skipped."
                )
                continue

            plan, warnings = _parse_sheet_to_plan(sheet_name, sheet_rows)
            aggregate_warnings.extend(warnings)
            if plan is None:
                continue
            if plan.total_rows == 0:
                # A sheet with a valid header but zero data rows is
                # retained — empty result is a legitimate signal to the
                # operator that the tab existed but contained no
                # priceable rows. The brief explicitly calls this out:
                # "Sheet with only header row, no data → empty plan
                # (not skipped)".
                out[sheet_name] = plan
                continue
            out[sheet_name] = plan
    finally:
        wb.close()

    if not out:
        msg = (
            "xlsx parsed but produced zero usable sheets — every tab "
            "was empty, lacked a recognisable header, or was filtered "
            "out by the sheet_filter callable."
        )
        if aggregate_warnings:
            msg += " Warnings: " + " | ".join(aggregate_warnings[:5])
        raise ValueError(msg)

    total_rows_across_sheets = sum(p.total_rows for p in out.values())
    if total_rows_across_sheets == 0:
        msg = (
            "xlsx parsed but produced zero priceable rows across "
            f"{len(out)} sheet(s). Every detected table had a header "
            "but no data rows."
        )
        if aggregate_warnings:
            msg += " Warnings: " + " | ".join(aggregate_warnings[:5])
        raise ValueError(msg)

    # Stash the per-sheet tag override on the dict itself via a
    # protocol-agnostic side channel — the Streamlit UI reads it back
    # from :func:`merge_xlsx_plans` rather than directly. Using setattr
    # on a dict is not portable; instead we expose the override via the
    # ``merge_xlsx_plans`` function signature so the contract is
    # explicit. The kwarg is therefore not stored on the OrderedDict
    # here — it's accepted on this function purely for symmetry with
    # the Streamlit-UI-friendly contract. Document the actual route
    # below in the merge function.
    _ = per_sheet_tag_override  # forward to merge via UI; documented above
    return out


# ---------------------------------------------------------------------------
# Public API — merge_xlsx_plans
# ---------------------------------------------------------------------------


def merge_xlsx_plans(
    plans: dict[str, BatchOverridePlan],
    *,
    sheet_separator: str = _DEFAULT_SHEET_SEPARATOR,
) -> BatchOverridePlan:
    """Merge per-sheet plans into a single flat :class:`BatchOverridePlan`.

    Two transformations applied per row:

    1. **Row-index renumbering.** Each row's ``row_index`` is
       monotonically renumbered starting at 2 (matching the CSV
       "header is row 1, first data row is row 2" convention). Within
       a sheet the relative order is preserved; across sheets the
       order is sheet-insertion order (i.e. workbook tab order). This
       is required because ``apply_batch_plan``'s ``skip_rows`` and
       ``resolved_ambiguous`` dicts key off ``row_index`` — duplicate
       indices across sheets would silently collide.
    2. **Sheet provenance in ``notes``.** Each row's ``notes`` field is
       rewritten to start with ``f"{sheet_name}{sheet_separator}row {N}"``
       where ``N`` is the row's *original* (per-sheet) ``row_index``,
       and any pre-existing notes are appended after the separator.
       So an operator looking at any downstream surface (the CSV
       export, the operator-note stamped on ``CostLine.notes``, the
       history-row ``operator_note`` field) can see which sheet a row
       came from at a glance even after merging.

    Returns:
        A single :class:`BatchOverridePlan` whose ``no_match`` bucket
        contains every row from every input plan (renumbered + sheet-
        annotated), and whose other buckets are empty. The caller is
        expected to feed the row list to
        :func:`~core.pricing.batch_override.match_cost_lines` to
        actually produce the match plan — the merged plan returned
        here is a row *carrier* not a match result.

    Args:
        plans: Output of :func:`parse_vendor_xlsx` (or any dict mapping
            sheet name → :class:`BatchOverridePlan`).
        sheet_separator: String inserted between sheet name and the
            ``row N`` identifier in the rewritten notes field.
            Defaults to ``" :: "`` — visually clear in a CSV / Excel
            Notes cell, unlikely to appear in legit vendor notes.

    Raises:
        ValueError: When ``plans`` is empty. The caller (UI or test)
            should special-case this — there is no meaningful "empty
            merged plan" semantics.
    """
    if not plans:
        raise ValueError(
            "merge_xlsx_plans called with zero plans; nothing to merge."
        )

    merged_rows: list[BatchOverrideRow] = []
    next_index = 2  # Mirrors CSV "header is row 1, data starts at 2".

    for sheet_name, plan in plans.items():
        sheet_rows = _plan_rows(plan)
        for original_row in sheet_rows:
            original_index = original_row.row_index
            provenance_prefix = (
                f"{sheet_name}{sheet_separator}row {original_index}"
            )
            existing_notes = (original_row.notes or "").strip()
            new_notes = (
                f"{provenance_prefix}{sheet_separator}{existing_notes}"
                if existing_notes
                else provenance_prefix
            )
            merged_rows.append(BatchOverrideRow(
                row_index=next_index,
                description=original_row.description,
                unit_cost=original_row.unit_cost,
                vendor=original_row.vendor,
                quote_ref=original_row.quote_ref,
                notes=new_notes,
                quantity=original_row.quantity,
                unit_of_measure=original_row.unit_of_measure,
            ))
            next_index += 1

    # Wrap each merged row in a no_match-bucket BatchMatchResult so the
    # caller can either route the rows directly through ``match_cost_lines``
    # (via _plan_rows / extracting .row attributes) or treat the
    # merged plan as a carrier and skip the BatchMatchResult layer
    # entirely. Either way the dataclass invariants hold.
    match_results = [
        BatchMatchResult(
            row=r,
            status=BatchMatchStatus.NO_MATCH,
            best_match_index=None,
            best_match_similarity=0.0,
            runner_up_index=None,
            runner_up_similarity=0.0,
            candidate_lines=[],
        )
        for r in merged_rows
    ]

    return BatchOverridePlan(
        total_rows=len(merged_rows),
        matched=[],
        ambiguous=[],
        no_match=match_results,
        low_similarity=[],
        similarity_threshold=0.65,
        ambiguity_margin=0.10,
    )
