"""Export the priced `Estimate` and supporting data to Excel and JSON."""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schemas import (
    AlternatePricingBasis,
    AlternateType,
    CostBand,
    CostSourceTier,
    Estimate,
    SheetExtraction,
)
from .takeoff import ProjectModel


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
DIVISION_FILL = PatternFill("solid", fgColor="D9E1F2")
SUPPRESSED_FILL = PatternFill("solid", fgColor="F2F2F2")  # light grey for unit-mismatch rows
SUPPRESSED_FONT = Font(italic=True, color="808080")

# Phase T6 — Hand-Takeoff Queue sheet header tint. Light orange / amber
# pulled from the PDF exporter's WARNING palette so the two surfaces
# read as the same "needs manual eyes" colour family. Exposed as a
# module-level constant so the exporter tests can pin the exact hex.
HAND_TAKEOFF_HEADER_FILL = PatternFill("solid", fgColor="FFC107")
HAND_TAKEOFF_HEADER_FONT = Font(bold=True, color="000000")
OPERATOR_REVIEW_HEADER_FILL = PatternFill("solid", fgColor="FFE066")
OPERATOR_REVIEW_HEADER_FONT = Font(bold=True, color="000000")


_BAND_LABELS: dict[CostBand, str] = {
    CostBand.AUTO_APPROVE: "AUTO",
    CostBand.OPERATOR_REVIEW: "REVIEW",
    CostBand.HAND_TAKEOFF: "HAND",
}


# Phase T7 — human-readable tier labels for Excel cells. The enum's
# string value (e.g. ``"exact_match"``) is fine for JSON / API consumers
# but a spreadsheet column reader expects Title Case. Exposed at module
# level so tests can pin the exact strings.
_TIER_LABELS: dict[CostSourceTier, str] = {
    CostSourceTier.EXACT_MATCH: "Exact Match",
    CostSourceTier.CATEGORY_MATCH: "Category Match",
    CostSourceTier.INTERPOLATED: "Interpolated",
    CostSourceTier.PARAMETRIC: "Parametric",
    CostSourceTier.MANUAL_OVERRIDE: "Manual Override",
    CostSourceTier.MISSING: "Missing",
}


# Phase T9.0 — bid-alternates worksheet header tint. Subtle blue-grey
# so the tab reads as a "supplementary cost detail" sheet alongside
# the operator / hand-takeoff queue sheets without competing visually
# with the headline Line Items + Summary tabs.
ALTERNATES_HEADER_FILL = PatternFill("solid", fgColor="A9C5E2")
ALTERNATES_HEADER_FONT = Font(bold=True, color="000000")

# Phase T9.0 — Title-Case labels for the bid-alternates sheet's
# Type and Pricing Basis columns. Mirrors the ``_TIER_LABELS`` /
# ``_BAND_LABELS`` pattern above. Surfaced as module-level constants
# so tests can pin exact strings.
_ALTERNATE_TYPE_LABELS: dict[AlternateType, str] = {
    AlternateType.ADDITIVE: "Additive",
    AlternateType.DEDUCTIVE: "Deductive",
    AlternateType.SUBSTITUTION: "Substitution",
    AlternateType.VE: "Value Engineering",
}

_ALTERNATE_BASIS_LABELS: dict[AlternatePricingBasis, str] = {
    AlternatePricingBasis.EXTRACTED_FROM_BID_FORM: "Extracted (bid form)",
    AlternatePricingBasis.OPERATOR_ENTERED: "Operator entered",
    AlternatePricingBasis.SYNTHESIZED_FROM_TAKEOFF: "Synthesized (takeoff)",
    AlternatePricingBasis.MISSING: "Missing — review",
}


def _alternate_type_label(atype) -> str:
    if isinstance(atype, AlternateType):
        return _ALTERNATE_TYPE_LABELS[atype]
    try:
        return _ALTERNATE_TYPE_LABELS[AlternateType(atype)]
    except Exception:
        return str(atype)


def _alternate_basis_label(basis) -> str:
    if isinstance(basis, AlternatePricingBasis):
        return _ALTERNATE_BASIS_LABELS[basis]
    try:
        return _ALTERNATE_BASIS_LABELS[AlternatePricingBasis(basis)]
    except Exception:
        return str(basis)


def _band_label(li) -> str:
    """Resolve the short ``AUTO/REVIEW/HAND`` label for a CostLine.

    Tolerates both the enum and the raw string value (Pydantic stores
    ``use_enum_values=False`` here, but persisted JSON round-trips can
    arrive with the string form depending on ``model_dump`` mode).
    """
    band = li.cost_band if hasattr(li, "cost_band") else CostBand.AUTO_APPROVE
    if isinstance(band, CostBand):
        return _BAND_LABELS[band]
    try:
        return _BAND_LABELS[CostBand(band)]
    except Exception:
        return str(band)


def _tier_label(li) -> str:
    """Resolve the Title-Case tier label for a CostLine (Phase T7).

    Mirrors :func:`_band_label` — tolerates either the enum form or the
    string form delivered by a JSON round-trip.
    """
    tier = getattr(li, "cost_source_tier", CostSourceTier.EXACT_MATCH)
    if isinstance(tier, CostSourceTier):
        return _TIER_LABELS[tier]
    try:
        return _TIER_LABELS[CostSourceTier(tier)]
    except Exception:
        return str(tier)


_SUPPORTING_DOC_KIND_PATTERNS: tuple[tuple[str, str], ...] = (
    # Order matters — first match wins.
    ("wage determination", "wage determination"),
    ("wage decision",      "wage determination"),
    ("davis-bacon",        "wage determination"),
    ("davis bacon",        "wage determination"),
    ("prevailing wage",    "wage determination"),
    ("wage rates",         "wage determination"),
    ("tax exemption",      "tax exemption certificate"),
    ("tax-exempt",         "tax exemption certificate"),
    ("sales tax",          "tax exemption certificate"),
    ("sample agreement",   "sample agreement"),
    ("sample contract",    "sample agreement"),
    ("sample construction services agreement", "sample agreement"),
    ("template agreement", "sample agreement"),
    ("csa template",       "sample agreement"),
    ("hub subcontracting plan", "HSP form"),
    ("subcontracting plan",     "HSP form"),
    ("hsp",                "HSP form"),
    ("ugsc",               "UGSC"),
    ("uniform general conditions",       "UGSC"),
    ("supplementary general conditions", "UGSC"),
    ("instructions to bidders",   "instructions to bidders"),
    ("instructions to offerors",  "instructions to bidders"),
    ("aia document",       "AIA template"),
)


def _classify_supporting_doc(p) -> str:
    """Map a supporting-document BidPackage to a short category label.

    Used by the Excel "Supporting Documents" sheet's Kind column.
    Falls back to "other" when no keyword matches.
    """
    haystack = " ".join([
        (p.pdf_name or "").lower(),
        (p.summary or "").lower(),
        (p.trade_name or "").lower(),
    ])
    for needle, label in _SUPPORTING_DOC_KIND_PATTERNS:
        if needle in haystack:
            return label
    return "other"


def _extract_supporting_doc_facts(p) -> tuple[str, str]:
    """Best-effort pull of (effective_date, wage_decision_number) from summary.

    Both are short strings or empty. We deliberately don't try to be clever —
    the LLM extraction puts these in `summary` when it found them; we just
    do a tiny regex scan as a convenience for the Excel reader.
    """
    import re as _re
    blob = " ".join([
        (p.summary or ""),
        (p.pdf_name or ""),
    ])
    effective_date = ""
    m = _re.search(
        r"(effective(?:\s+date)?[:\s]+)([A-Z][a-z]+\s+\d{1,2},\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})",
        blob,
        flags=_re.IGNORECASE,
    )
    if m:
        effective_date = m.group(2)
    wage_decision = ""
    m = _re.search(r"\b(TX\s?\d{8}|WD\s*\d{6,}|\d{4}-\d{4,})\b", blob)
    if m:
        wage_decision = m.group(1).replace(" ", "")
    return effective_date, wage_decision


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        longest = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max(longest + 2, 10), max_width)


def _write_header(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


# Column schema shared by the two Phase T6 queue sheets. Kept as a
# module-level tuple so tests can reference it and any future column
# addition lands in one place. Phase T7 appended ``Price Confidence``,
# ``Cost Source Tier`` and ``Combined Confidence`` after ``Confidence``
# so a reviewer scanning either queue sees the catalog-completeness
# axis right next to the qty-side confidence.
_QUEUE_SHEET_HEADERS: tuple[str, ...] = (
    "CSI Code",
    "Description",
    "Qty",
    "Unit",
    "Unit Cost",
    "Total Cost",
    "Confidence",
    "Price Confidence",
    "Cost Source Tier",
    "Combined Confidence",
    "Source",
    "Notes",
)


def _render_queue_sheet(
    wb,
    *,
    title: str,
    lines: list,
    header_fill: PatternFill,
    header_font: Font,
    empty_note: str,
) -> None:
    """Render one of the Phase T6 queue sheets (Operator-Review / Hand-Takeoff).

    Always creates the worksheet even when ``lines`` is empty so callers
    + tests can rely on the tab existing. The header row carries the
    queue-specific tint so the action-required signal travels with the
    workbook regardless of theme.

    ``Source`` column shows comma-joined ``source_sheet_ids`` (typically
    the originating drawing-sheet IDs) — this is what an estimator scans
    when deciding which PDF page to flip to for a manual count.
    """
    ws = wb.create_sheet(title)
    for i, h in enumerate(_QUEUE_SHEET_HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    if not lines:
        # Friendly empty-state note in A2 so the reader knows the tab is
        # intentionally empty (not a broken render).
        ws.cell(row=2, column=1, value=empty_note).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.freeze_panes = "A2"
        _autosize(ws, max_width=80)
        return

    for r_idx, li in enumerate(lines, start=2):
        ws.cell(row=r_idx, column=1, value=li.csi_section or li.csi_division)
        ws.cell(row=r_idx, column=2, value=li.description).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row=r_idx, column=3, value=li.quantity).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=4, value=li.unit)
        ws.cell(row=r_idx, column=5, value=li.unit_cost).number_format = "$#,##0.00"
        ws.cell(row=r_idx, column=6, value=li.total_cost).number_format = "$#,##0.00"
        ws.cell(row=r_idx, column=7, value=li.confidence).number_format = "0.00"
        # Phase T7 — price-side axis next to the qty-side ``Confidence`` cell.
        ws.cell(row=r_idx, column=8, value=getattr(li, "price_confidence", 1.0)).number_format = "0.00"
        ws.cell(row=r_idx, column=9, value=_tier_label(li))
        combined = li.combined_confidence if hasattr(li, "combined_confidence") else li.confidence
        ws.cell(row=r_idx, column=10, value=combined).number_format = "0.00"
        ws.cell(row=r_idx, column=11, value=", ".join(li.source_sheet_ids))
        ws.cell(row=r_idx, column=12, value=li.notes or "").alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    ws.freeze_panes = "A2"
    _autosize(ws, max_width=80)


# Phase T9.0 — Bid Alternates worksheet column schema. Module-level so
# tests can reference the exact contract.
_ALTERNATES_SHEET_HEADERS: tuple[str, ...] = (
    "Alternate ID",
    "Type",
    "Description",
    "Cost Delta",
    "Pricing Basis",
    "Confidence",
    "Included in Base",
    "Bid Package",
    "Source Sheet",
    "Related CSI",
    "Operator Notes",
)


def _render_alternates_sheet(wb, estimate: Estimate) -> None:
    """Phase T9.0 — render the "Bid Alternates" worksheet.

    Always called from :func:`export_estimate_xlsx` but bails out early
    (creates no sheet) when the estimate has zero priced alternates —
    keeps the standard exporter output clean for projects without any
    alternates surfaced (most non-government bid sets).

    Footer rows below the line items:

      * Base bid (no alternates)
      * Base + all additive
      * Base + all deductive (incl. VE)
      * Base + default-selected (the alternates marked
        ``included_by_default=True``)

    All currency cells use ``$#,##0.00``; confidence cells use ``0.00``.
    """
    alts = list(getattr(estimate, "alternates", None) or [])
    if not alts:
        return

    ws = wb.create_sheet("Bid Alternates")
    headers = list(_ALTERNATES_SHEET_HEADERS)
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = ALTERNATES_HEADER_FILL
        cell.font = ALTERNATES_HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for r_idx, a in enumerate(alts, start=2):
        ws.cell(row=r_idx, column=1, value=a.alternate_id)
        ws.cell(row=r_idx, column=2, value=_alternate_type_label(a.alternate_type))
        ws.cell(row=r_idx, column=3, value=a.description).alignment = Alignment(wrap_text=True)
        if a.cost_delta is not None:
            ws.cell(row=r_idx, column=4, value=a.cost_delta).number_format = "$#,##0.00"
        else:
            ws.cell(row=r_idx, column=4, value="—")
        ws.cell(row=r_idx, column=5, value=_alternate_basis_label(a.pricing_basis))
        ws.cell(row=r_idx, column=6, value=a.confidence).number_format = "0.00"
        ws.cell(row=r_idx, column=7, value="YES" if a.included_in_base else "")
        ws.cell(row=r_idx, column=8, value=a.bid_package_id or "")
        ws.cell(row=r_idx, column=9, value=a.source_sheet or "")
        ws.cell(row=r_idx, column=10, value=", ".join(a.related_csi or []))
        ws.cell(row=r_idx, column=11, value=a.operator_notes or "").alignment = Alignment(
            wrap_text=True
        )

    # Footer rows: base bid + alternate-selection rollups.
    footer_row = len(alts) + 3
    additive_total = estimate.alternates_total_additive
    deductive_total = estimate.alternates_total_deductive
    substitution_total = estimate.alternates_total_substitution

    additive_ids = {
        a.alternate_id for a in alts
        if a.alternate_type == AlternateType.ADDITIVE and a.cost_delta is not None
    }
    deductive_ids = {
        a.alternate_id for a in alts
        if a.alternate_type in (AlternateType.DEDUCTIVE, AlternateType.VE)
        and a.cost_delta is not None
    }
    substitution_ids = {
        a.alternate_id for a in alts
        if a.alternate_type == AlternateType.SUBSTITUTION and a.cost_delta is not None
    }
    default_ids = set(estimate.alternates_selected_default or set())

    rows: list[tuple[str, float | None]] = [
        ("Base bid (no alternates)", estimate.subtotal_base_only),
        (
            "Base + all additive alternates",
            estimate.subtotal_with_alternates_selected(additive_ids),
        ),
        (
            "Base + all deductive / VE alternates",
            estimate.subtotal_with_alternates_selected(deductive_ids),
        ),
        (
            "Base + all substitution alternates",
            estimate.subtotal_with_alternates_selected(substitution_ids),
        ),
        (
            "Base + default-selected alternates",
            estimate.subtotal_with_alternates_selected(default_ids),
        ),
    ]
    for offset, (label, value) in enumerate(rows):
        r = footer_row + offset
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True)
        c2 = ws.cell(row=r, column=4, value=round(value, 2))
        c2.number_format = "$#,##0.00"
        c2.font = Font(bold=True)

    # Aggregate counts (informational; clearly distinct from the footer
    # subtotals above).
    counts_row = footer_row + len(rows) + 2
    summary_pairs: list[tuple[str, object]] = [
        ("Additive alternates total", additive_total),
        ("Deductive / VE alternates total", deductive_total),
        ("Substitution alternates total", substitution_total),
        ("Alternates missing a price", estimate.alternates_count_missing),
        ("Total alternates", len(alts)),
    ]
    for offset, (label, value) in enumerate(summary_pairs):
        r = counts_row + offset
        ws.cell(row=r, column=1, value=label).font = Font(italic=True)
        cell = ws.cell(row=r, column=4, value=value)
        if isinstance(value, float):
            cell.number_format = "$#,##0.00"
        else:
            cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    _autosize(ws, max_width=80)


def export_estimate_xlsx(
    estimate: Estimate,
    project: ProjectModel,
    csi_titles: dict[str, str],
    extractions: Iterable[SheetExtraction] | None = None,
) -> bytes:
    wb = Workbook()

    # ----- Summary -----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Project"
    ws["B1"] = estimate.project_name
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Region multiplier"
    ws["B3"] = estimate.region_multiplier

    ws["A5"] = "Subtotal";    ws["B5"] = estimate.subtotal;     ws["B5"].number_format = "$#,##0.00"
    ws["A6"] = f"Contingency ({estimate.contingency_pct:.1f}%)"
    ws["B6"] = estimate.contingency; ws["B6"].number_format = "$#,##0.00"
    ws["A7"] = f"Overhead ({estimate.overhead_pct:.1f}%)"
    ws["B7"] = estimate.overhead;    ws["B7"].number_format = "$#,##0.00"
    ws["A8"] = f"Profit ({estimate.profit_pct:.1f}%)"
    ws["B8"] = estimate.profit;      ws["B8"].number_format = "$#,##0.00"
    ws["A9"] = "GRAND TOTAL"
    ws["A9"].font = Font(bold=True, size=12)
    ws["B9"] = estimate.grand_total
    ws["B9"].number_format = "$#,##0.00"
    ws["B9"].font = Font(bold=True, size=12)

    # Phase T6 — band-aware totals + queue counts. Placed immediately
    # under the headline grand total so an estimator reviewing the
    # Project Summary sheet sees the breakdown next to the number it
    # explains.
    ws["A10"] = "Auto-Approve Total"
    ws["B10"] = estimate.total_auto_approve;       ws["B10"].number_format = "$#,##0.00"
    ws["A11"] = "Operator-Review Total"
    ws["B11"] = estimate.total_operator_review;    ws["B11"].number_format = "$#,##0.00"
    ws["A12"] = "Hand-Takeoff Total (informational, not in grand total)"
    ws["B12"] = estimate.total_hand_takeoff;       ws["B12"].number_format = "$#,##0.00"
    ws["A13"] = "Grand Total (Auto-Only)"
    ws["B13"] = estimate.grand_total_auto_only;    ws["B13"].number_format = "$#,##0.00"
    ws["A14"] = "Grand Total (Auto + Review)"
    ws["B14"] = estimate.grand_total_with_review;  ws["B14"].number_format = "$#,##0.00"
    ws["A15"] = f"Lines Needing Manual Takeoff: {estimate.hand_takeoff_count}"
    ws["A16"] = f"Lines Needing Operator Review: {estimate.operator_review_count}"

    # Phase T9.0 — bid-alternates summary block. Surfaces the count
    # + the total additive / deductive deltas next to the headline
    # subtotal so a reviewer immediately sees the upside / downside
    # range without flipping to the "Bid Alternates" tab. Renders
    # only when the estimate carries at least one alternate so the
    # default exporter output (Project Summary row layout) is
    # unchanged for projects without any.
    alt_count = len(getattr(estimate, "alternates", None) or [])
    if alt_count:
        add_total = estimate.alternates_total_additive
        ded_total = estimate.alternates_total_deductive
        sub_total = estimate.alternates_total_substitution
        missing = estimate.alternates_count_missing
        ws["A17"] = "Bid Alternates"
        ws["A17"].font = Font(bold=True)
        ws["A18"] = f"Alternates: {alt_count} ({missing} missing a price)"
        ws["A19"] = "Total additive potential"
        ws["B19"] = add_total
        ws["B19"].number_format = "$#,##0.00"
        ws["A20"] = "Total deductive / VE potential"
        ws["B20"] = ded_total
        ws["B20"].number_format = "$#,##0.00"
        ws["A21"] = "Total substitution potential (net)"
        ws["B21"] = sub_total
        ws["B21"].number_format = "$#,##0.00"
        division_anchor_row = 23
    else:
        division_anchor_row = 18

    ws.cell(row=division_anchor_row, column=1, value="By CSI division").font = Font(bold=True)
    _write_header(ws, division_anchor_row + 1, ["Div", "Title", "Subtotal"])
    row = division_anchor_row + 2
    for div in sorted(estimate.by_division):
        ws.cell(row=row, column=1, value=div)
        ws.cell(row=row, column=2, value=csi_titles.get(div, ""))
        c = ws.cell(row=row, column=3, value=estimate.by_division[div])
        c.number_format = "$#,##0.00"
        row += 1

    # F3 — drawing-prepass coverage. Single-row tile under the division
    # breakdown so the reader can see, at a glance, how much of the run
    # came out of the deterministic pre-pass vs the vision LLM.
    extraction_list = list(extractions or [])
    drawing_extractions = [
        e for e in extraction_list
        if e.bid_package is None and (e.prepass is not None or e.lm_skipped or e.summary)
    ]
    if drawing_extractions:
        row += 1
        skipped = sum(1 for e in drawing_extractions if e.lm_skipped)
        total = len(drawing_extractions)
        skipped_pct = (skipped / total * 100.0) if total else 0.0
        ws.cell(row=row, column=1, value="Prepass coverage").font = Font(bold=True)
        ws.cell(
            row=row, column=2,
            value=(
                f"{skipped_pct:.0f}% pages prepass-only / "
                f"{100.0 - skipped_pct:.0f}% pages LLM-augmented "
                f"({skipped} of {total})"
            ),
        )

    # By cost category rollup (labor / material / equipment / sub / other)
    row += 1
    by_cat = estimate.by_cost_category
    if by_cat:
        ws.cell(row=row, column=1, value="By cost category").font = Font(bold=True)
        row += 1
        _write_header(ws, row, ["Category", "Subtotal", "% of subtotal"])
        row += 1
        subtotal = estimate.subtotal or 0.0
        for cat, total in sorted(by_cat.items(), key=lambda kv: -kv[1]):
            ws.cell(row=row, column=1, value=cat.title())
            c_total = ws.cell(row=row, column=2, value=total)
            c_total.number_format = "$#,##0.00"
            pct = (total / subtotal * 100.0) if subtotal else 0.0
            c_pct = ws.cell(row=row, column=3, value=round(pct, 2))
            c_pct.number_format = "0.00\"%\""
            row += 1

    # Phase T7 — Cost Source Tier Breakdown. One row per
    # :class:`CostSourceTier` in enum order so the table reads top-to-
    # bottom as "best → worst" catalog quality. Always emitted (even on
    # a fully-empty estimate) so downstream parsers can rely on the
    # block existing. The percentage denominator is the headline
    # subtotal (AUTO + REVIEW; excludes HAND_TAKEOFF / suppressed)
    # which matches the PDF subscript denominator and reconciles back
    # to the same number the reader sees on the headline tile.
    row += 1
    ws.cell(row=row, column=1, value="Cost Source Tier Breakdown").font = Font(bold=True)
    row += 1
    _write_header(ws, row, ["Tier", "Line Count", "Total $", "% of Subtotal"])
    row += 1
    tier_counts = estimate.count_by_tier
    tier_totals = estimate.total_by_tier
    subtotal_for_tier = estimate.subtotal or 0.0
    total_lines_all = 0
    total_dollars_all = 0.0
    for tier in CostSourceTier:
        cnt = tier_counts.get(tier, 0)
        amount = tier_totals.get(tier, 0.0)
        ws.cell(row=row, column=1, value=_TIER_LABELS[tier])
        ws.cell(row=row, column=2, value=cnt)
        c_total = ws.cell(row=row, column=3, value=amount)
        c_total.number_format = "$#,##0.00"
        pct = (amount / subtotal_for_tier * 100.0) if subtotal_for_tier else 0.0
        c_pct = ws.cell(row=row, column=4, value=round(pct, 2))
        c_pct.number_format = "0.00\"%\""
        total_lines_all += cnt
        total_dollars_all += amount
        row += 1
    # Reconciliation row matching the per-tier table — sums to the
    # full ``len(line_items)`` and to the sum of ``total_cost`` across
    # every tier (which is NOT identical to ``estimate.subtotal`` since
    # it includes MISSING / HAND_TAKEOFF dollars at $0; the row makes
    # the reconciliation explicit).
    ws.cell(row=row, column=1, value=f"Total: {total_lines_all} lines").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_lines_all).font = Font(bold=True)
    c_total = ws.cell(row=row, column=3, value=round(total_dollars_all, 2))
    c_total.number_format = "$#,##0.00"
    c_total.font = Font(bold=True)
    row += 1

    _autosize(ws)

    # ----- Line items (a.k.a. "Cost Estimate") -----
    ws = wb.create_sheet("Line Items")
    headers = [
        "Div", "CSI Section", "Category", "Description",
        "Raw Qty", "Waste", "Quantity", "Unit",
        "Unit Cost", "Total", "Band", "Suppressed",
        "Confidence",
        # Phase T7 — three new columns inserted immediately after the
        # qty-side ``Confidence`` column so a reviewer sees both
        # confidence axes side by side. The order
        # (Confidence → Price Confidence → Cost Source Tier →
        # Combined Confidence) is fixed by the Phase T7 brief.
        "Price Confidence", "Cost Source Tier", "Combined Confidence",
        "Source Sheets",
        "Cost Source", "CWICR Similarity", "Cost-DB Key", "Notes",
    ]
    _write_header(ws, 1, headers)

    last_div = None
    row = 2
    for li in estimate.line_items:
        if li.csi_division != last_div:
            ws.cell(row=row, column=1, value=f"Division {li.csi_division} - {csi_titles.get(li.csi_division, '')}").fill = DIVISION_FILL
            for col in range(2, len(headers) + 1):
                ws.cell(row=row, column=col).fill = DIVISION_FILL
            ws.cell(row=row, column=1).font = Font(bold=True)
            last_div = li.csi_division
            row += 1

        cat_val = li.cost_category.value if hasattr(li.cost_category, "value") else str(li.cost_category)
        # Decode cost source family + CWICR row-id / similarity for the
        # post-F1 "Cost Source" + "CWICR Similarity" columns. We keep the
        # raw cost-DB key around in its own column so users can still see
        # the exact lookup key the seed DB returned.
        raw_src = li.cost_source or ""
        if raw_src.startswith("cwicr:"):
            src_family = "cwicr"
            cwicr_sim = li.confidence
        elif raw_src in ("", "(no match)"):
            src_family = "no match"
            cwicr_sim = None
        else:
            src_family = "seed"
            cwicr_sim = None

        ws.cell(row=row, column=1, value=li.csi_division)
        ws.cell(row=row, column=2, value=li.csi_section or "")
        ws.cell(row=row, column=3, value=cat_val)
        ws.cell(row=row, column=4, value=li.description)
        ws.cell(row=row, column=5, value=li.raw_quantity if li.raw_quantity is not None else li.quantity).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=li.waste_factor).number_format = "0.00"
        ws.cell(row=row, column=7, value=li.quantity).number_format = "#,##0.00"
        ws.cell(row=row, column=8, value=li.unit)
        ws.cell(row=row, column=9, value=li.unit_cost).number_format = "$#,##0.00"
        ws.cell(row=row, column=10, value=li.total_cost).number_format = "$#,##0.00"
        ws.cell(row=row, column=11, value=_band_label(li))
        ws.cell(row=row, column=12, value="YES" if li.suppressed else "")
        ws.cell(row=row, column=13, value=li.confidence).number_format = "0.00"
        # Phase T7 columns — same numeric format as ``Confidence`` for
        # the price-side and combined values; tier is a string label.
        ws.cell(row=row, column=14, value=getattr(li, "price_confidence", 1.0)).number_format = "0.00"
        ws.cell(row=row, column=15, value=_tier_label(li))
        combined = li.combined_confidence if hasattr(li, "combined_confidence") else li.confidence
        ws.cell(row=row, column=16, value=combined).number_format = "0.00"
        ws.cell(row=row, column=17, value=", ".join(li.source_sheet_ids))
        ws.cell(row=row, column=18, value=src_family)
        sim_cell = ws.cell(row=row, column=19, value=cwicr_sim)
        if cwicr_sim is not None:
            sim_cell.number_format = "0.00"
        ws.cell(row=row, column=20, value=raw_src)
        ws.cell(row=row, column=21, value=li.notes or "")

        if li.suppressed:
            # Grey-shade the entire row and italicize so the reader sees at a
            # glance that this line did NOT roll into the totals below.
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = SUPPRESSED_FILL
                cell.font = SUPPRESSED_FONT
        row += 1

    ws.freeze_panes = "A2"
    _autosize(ws)

    # ----- Operator Review Queue + Hand Takeoff Queue (Phase T6) -----
    # Two parallel sheets, identical schema. Both are ALWAYS created (even
    # when the queue is empty) so downstream consumers can rely on the
    # tabs existing. Rendered after the main Line Items sheet so the
    # reader's eye flows: full estimate → review queue → manual-takeoff
    # worklist.
    _render_queue_sheet(
        wb,
        title="Operator Review Queue",
        lines=list(estimate.operator_review_line_items),
        header_fill=OPERATOR_REVIEW_HEADER_FILL,
        header_font=OPERATOR_REVIEW_HEADER_FONT,
        empty_note="No operator-review lines — all priced rows cleared the "
                   "0.85 auto-approve threshold.",
    )
    _render_queue_sheet(
        wb,
        title="Hand Takeoff Queue",
        # HAND queue surfaces BOTH low-confidence priced lines and any
        # suppressed (unit-mismatch) lines, since both need manual eyes.
        # ``Estimate.hand_takeoff_line_items`` already filters to the
        # HAND_TAKEOFF band regardless of suppression.
        lines=list(estimate.hand_takeoff_line_items),
        header_fill=HAND_TAKEOFF_HEADER_FILL,
        header_font=HAND_TAKEOFF_HEADER_FONT,
        empty_note="No hand-takeoff lines — every row had ≥ 0.65 confidence "
                   "and no unit-mismatch suppression triggered.",
    )

    # ----- Bid Alternates (Phase T9.0) -----
    # Renders only when the estimate carries at least one priced
    # alternate. Sheet is omitted entirely (no "(empty)" stub) for
    # projects without alternates — keeps the standard exporter
    # output unchanged for those cases. Placed after the queue sheets
    # because alternates are a supplementary cost surface, not part
    # of the headline line-item flow.
    _render_alternates_sheet(wb, estimate)

    # ----- Rooms -----
    if project.rooms:
        ws = wb.create_sheet("Rooms")
        headers = ["Number", "Name", "Area (SF)", "Perimeter (LF)", "Ceiling (FT)",
                   "Floor", "Base", "Walls", "Ceiling", "Source"]
        _write_header(ws, 1, headers)
        for i, r in enumerate(project.rooms, start=2):
            ws.cell(row=i, column=1, value=r.number or "")
            ws.cell(row=i, column=2, value=r.name)
            ws.cell(row=i, column=3, value=r.area_sqft).number_format = "#,##0.00"
            ws.cell(row=i, column=4, value=r.perimeter_ft).number_format = "#,##0.00"
            ws.cell(row=i, column=5, value=r.ceiling_height_ft)
            ws.cell(row=i, column=6, value=r.floor_finish or "")
            ws.cell(row=i, column=7, value=r.base_finish or "")
            ws.cell(row=i, column=8, value=r.wall_finish or "")
            ws.cell(row=i, column=9, value=r.ceiling_finish or "")
            ws.cell(row=i, column=10, value=r.source_sheet_id or "")
        ws.freeze_panes = "A2"
        _autosize(ws)

    # ----- Doors / Windows -----
    if project.doors:
        ws = wb.create_sheet("Doors")
        _write_header(ws, 1, ["Mark", "Type", "W (in)", "H (in)", "Rating", "Hardware", "Notes", "Source"])
        for i, d in enumerate(project.doors, start=2):
            ws.cell(row=i, column=1, value=d.mark)
            ws.cell(row=i, column=2, value=d.type or "")
            ws.cell(row=i, column=3, value=d.width_in)
            ws.cell(row=i, column=4, value=d.height_in)
            ws.cell(row=i, column=5, value=d.rating or "")
            ws.cell(row=i, column=6, value=d.hardware_set or "")
            ws.cell(row=i, column=7, value=d.notes or "")
            ws.cell(row=i, column=8, value=d.source_sheet_id or "")
        ws.freeze_panes = "A2"
        _autosize(ws)

    if project.windows:
        ws = wb.create_sheet("Windows")
        _write_header(ws, 1, ["Mark", "Type", "W (in)", "H (in)", "Glazing", "Notes", "Source"])
        for i, w in enumerate(project.windows, start=2):
            ws.cell(row=i, column=1, value=w.mark)
            ws.cell(row=i, column=2, value=w.type or "")
            ws.cell(row=i, column=3, value=w.width_in)
            ws.cell(row=i, column=4, value=w.height_in)
            ws.cell(row=i, column=5, value=w.glazing or "")
            ws.cell(row=i, column=6, value=w.notes or "")
            ws.cell(row=i, column=7, value=w.source_sheet_id or "")
        ws.freeze_panes = "A2"
        _autosize(ws)

    # ----- Project Info -----
    pi = project.project_info
    if pi and (pi.name or pi.number):
        ws = wb.create_sheet("Project Info", 0)
        ws["A1"] = "Project Information"
        ws["A1"].font = Font(bold=True, size=14)
        rows = [
            ("Project name",     pi.name or ""),
            ("Project number",   pi.number or ""),
            ("Location",         pi.location or ""),
            ("Owner",            pi.owner or ""),
            ("Contractor / GC",  pi.contractor or ""),
            ("Bid due",          pi.bid_due or ""),
            ("Source PDFs",      ", ".join(pi.sources or [])),
        ]
        for i, (k, v) in enumerate(rows, start=3):
            c1 = ws.cell(row=i, column=1, value=k); c1.font = Font(bold=True)
            ws.cell(row=i, column=2, value=v)
        _autosize(ws, max_width=80)

    # ----- Bid Packages (trade packages only) -----
    trade_packages = [
        p for p in project.bid_packages if p.document_kind == "trade_package"
    ]
    supporting_docs = [
        p for p in project.bid_packages if p.document_kind == "supporting_document"
    ]
    ws = wb.create_sheet("Bid Packages")
    headers = [
        "Pkg #", "Trade", "PDF", "Owner", "General Contractor",
        "Project", "Project #",
        "CSI Divs", "CSI Sections",
        "# Inclusions", "# Exclusions", "# Alternates", "# Unit Prices",
        "Refd Drawings", "Refd Specs", "Summary",
    ]
    _write_header(ws, 1, headers)
    if not trade_packages:
        ws.cell(
            row=2,
            column=1,
            value="No trade bid packages in this project.",
        ).alignment = Alignment(wrap_text=True, vertical="top")
    else:
        for i, p in enumerate(sorted(trade_packages, key=lambda x: x.package_number or x.pdf_name), start=2):
            ws.cell(row=i, column=1, value=p.package_number or "")
            ws.cell(row=i, column=2, value=p.trade_name or "")
            ws.cell(row=i, column=3, value=p.pdf_name)
            ws.cell(row=i, column=4, value=p.owner or "")
            ws.cell(row=i, column=5, value=p.gc or "")
            ws.cell(row=i, column=6, value=p.project_name or "")
            ws.cell(row=i, column=7, value=p.project_number or "")
            ws.cell(row=i, column=8, value=", ".join(p.csi_divisions))
            ws.cell(row=i, column=9, value=", ".join(p.csi_sections))
            ws.cell(row=i, column=10, value=len(p.inclusions))
            ws.cell(row=i, column=11, value=len(p.exclusions))
            ws.cell(row=i, column=12, value=len(p.alternates))
            ws.cell(row=i, column=13, value=len(p.unit_prices))
            ws.cell(row=i, column=14, value=", ".join(p.referenced_drawings))
            ws.cell(row=i, column=15, value=", ".join(p.referenced_specs))
            ws.cell(row=i, column=16, value=p.summary or "").alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    _autosize(ws, max_width=80)

    # Detail sheet: one row per inclusion/exclusion/alternate/unit price
    ws = wb.create_sheet("Scope Matrix")
    _write_header(ws, 1, ["Pkg #", "Trade", "Type", "Detail", "Add/Deduct", "Unit", "Amount"])
    if not trade_packages:
        ws.cell(
            row=2,
            column=1,
            value="No scope rows — no trade bid packages in this project.",
        ).alignment = Alignment(wrap_text=True, vertical="top")
    else:
        row = 2
        for p in sorted(trade_packages, key=lambda x: x.package_number or x.pdf_name):
            for inc in p.inclusions:
                ws.cell(row=row, column=1, value=p.package_number or "")
                ws.cell(row=row, column=2, value=p.trade_name or "")
                ws.cell(row=row, column=3, value="Inclusion")
                ws.cell(row=row, column=4, value=inc).alignment = Alignment(wrap_text=True)
                row += 1
            for exc in p.exclusions:
                ws.cell(row=row, column=1, value=p.package_number or "")
                ws.cell(row=row, column=2, value=p.trade_name or "")
                ws.cell(row=row, column=3, value="Exclusion")
                ws.cell(row=row, column=4, value=exc).alignment = Alignment(wrap_text=True)
                row += 1
            for a in p.alternates:
                ws.cell(row=row, column=1, value=p.package_number or "")
                ws.cell(row=row, column=2, value=p.trade_name or "")
                ws.cell(row=row, column=3, value=f"Alternate {a.number or ''}".strip())
                ws.cell(row=row, column=4, value=a.description).alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=5, value=a.add_or_deduct or "")
                if a.amount is not None:
                    ws.cell(row=row, column=7, value=a.amount).number_format = "$#,##0.00"
                row += 1
            for u in p.unit_prices:
                ws.cell(row=row, column=1, value=p.package_number or "")
                ws.cell(row=row, column=2, value=p.trade_name or "")
                ws.cell(row=row, column=3, value="Unit Price")
                ws.cell(row=row, column=4, value=u.description).alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=6, value=u.unit or "")
                if u.amount is not None:
                    ws.cell(row=row, column=7, value=u.amount).number_format = "$#,##0.00"
                row += 1
    ws.freeze_panes = "A2"
    _autosize(ws, max_width=100)

    # ----- Supporting Documents -----
    # Wage determinations, sample agreements, tax-exemption certificates,
    # HSP templates, UGSC docs — anything `document_kind == supporting_document`.
    # Disjoint from "Bid Packages" above; the master `bid_packages` list
    # keeps both so reconcile / aggregation still see them.
    if supporting_docs:
        ws = wb.create_sheet("Supporting Documents")
        _write_header(
            ws, 1,
            ["Filename", "Source Page", "Kind", "Effective Date",
             "Wage Decision #", "Notes"],
        )
        for i, p in enumerate(sorted(supporting_docs, key=lambda x: x.pdf_name), start=2):
            kind = _classify_supporting_doc(p)
            effective_date, wage_decision = _extract_supporting_doc_facts(p)
            ws.cell(row=i, column=1, value=p.pdf_name)
            ws.cell(row=i, column=2, value="")  # page reference — not currently tracked
            ws.cell(row=i, column=3, value=kind)
            ws.cell(row=i, column=4, value=effective_date)
            ws.cell(row=i, column=5, value=wage_decision)
            ws.cell(row=i, column=6, value=p.summary or "").alignment = Alignment(wrap_text=True)
        ws.freeze_panes = "A2"
        _autosize(ws, max_width=100)

    # ----- Scope Coverage (aggregated inclusions / exclusions) -----
    agg_inc = getattr(project, "aggregated_inclusions", []) or []
    agg_exc = getattr(project, "aggregated_exclusions", []) or []
    if agg_inc or agg_exc:
        ws = wb.create_sheet("Scope Coverage")
        ws["A1"] = "Aggregated scope coverage"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = (
            "Deduplicated across all bid packages. 'Source packages' shows "
            "every bid-package PDF whose inclusion/exclusion list contributed "
            "to this line."
        )
        ws["A2"].alignment = Alignment(wrap_text=True)

        row = 4
        ws.cell(row=row, column=1, value=f"Inclusions ({len(agg_inc)})").font = Font(bold=True, size=12)
        row += 1
        _write_header(ws, row, ["#", "Inclusion", "# Packages", "Source Packages"])
        row += 1
        for i, item in enumerate(agg_inc, start=1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=item.text).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=3, value=len(item.source_packages))
            ws.cell(row=row, column=4, value=", ".join(item.source_packages)).alignment = Alignment(wrap_text=True)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value=f"Exclusions ({len(agg_exc)})").font = Font(bold=True, size=12)
        row += 1
        _write_header(ws, row, ["#", "Exclusion", "# Packages", "Source Packages"])
        row += 1
        for i, item in enumerate(agg_exc, start=1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=item.text).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=3, value=len(item.source_packages))
            ws.cell(row=row, column=4, value=", ".join(item.source_packages)).alignment = Alignment(wrap_text=True)
            row += 1

        _autosize(ws, max_width=100)

    # ----- Sheets analysed -----
    if project.sheet_summaries:
        ws = wb.create_sheet("Sheets")
        _write_header(ws, 1, ["Sheet", "LLM skipped", "Prepass conf", "Summary"])
        # Index extractions by sheet_id for the LLM-skipped column.
        skipped_by_sid: dict[str, tuple[bool, float | None]] = {}
        for ex in extraction_list:
            conf = ex.prepass.confidence if ex.prepass is not None else None
            skipped_by_sid[ex.sheet_id] = (bool(ex.lm_skipped), conf)
        for i, (sid, summary) in enumerate(sorted(project.sheet_summaries.items()), start=2):
            ws.cell(row=i, column=1, value=sid)
            lm_skipped, conf = skipped_by_sid.get(sid, (False, None))
            ws.cell(row=i, column=2, value="YES" if lm_skipped else "")
            if conf is not None:
                c = ws.cell(row=i, column=3, value=conf)
                c.number_format = "0.00"
            ws.cell(row=i, column=4, value=summary).alignment = Alignment(wrap_text=True)
        ws.freeze_panes = "A2"
        _autosize(ws, max_width=120)

    # ----- Warnings -----
    if project.warnings:
        ws = wb.create_sheet("Warnings")
        _write_header(ws, 1, ["#", "Warning"])
        for i, w in enumerate(project.warnings, start=2):
            ws.cell(row=i, column=1, value=i - 1)
            ws.cell(row=i, column=2, value=w).alignment = Alignment(wrap_text=True)
        _autosize(ws, max_width=120)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def export_estimate_json(estimate: Estimate, project: ProjectModel) -> str:
    pi = project.project_info
    payload = {
        "project_name": estimate.project_name,
        "project_info": {
            "name":       pi.name,
            "number":     pi.number,
            "location":   pi.location,
            "owner":      pi.owner,
            "contractor": pi.contractor,
            "bid_due":    pi.bid_due,
            "sources":    pi.sources or [],
        },
        "region_multiplier": estimate.region_multiplier,
        "subtotal": estimate.subtotal,
        "contingency_pct": estimate.contingency_pct,
        "contingency": estimate.contingency,
        "overhead_pct": estimate.overhead_pct,
        "overhead": estimate.overhead,
        "profit_pct": estimate.profit_pct,
        "profit": estimate.profit,
        "grand_total": estimate.grand_total,
        # Phase T6 — band-aware aggregates surfaced alongside the headline
        # number so JSON consumers (downstream proposal builders, the CLI
        # `--client-pdf` flag, the Streamlit Estimate tab) can render the
        # AUTO / REVIEW / HAND breakdown without re-walking line_items.
        "grand_total_with_review": estimate.grand_total_with_review,
        "grand_total_auto_only": estimate.grand_total_auto_only,
        "total_auto_approve": estimate.total_auto_approve,
        "total_operator_review": estimate.total_operator_review,
        "total_hand_takeoff": estimate.total_hand_takeoff,
        "auto_approve_count": estimate.auto_approve_count,
        "operator_review_count": estimate.operator_review_count,
        "hand_takeoff_count": estimate.hand_takeoff_count,
        # Phase T7 — catalog-completeness aggregates. Keys are the
        # CostSourceTier string values (``exact_match`` / ``category_match``
        # / ``interpolated`` / ``parametric`` / ``manual_override`` /
        # ``missing``) so JSON consumers can render the breakdown
        # without re-walking ``line_items``.
        "total_by_tier": {
            tier.value: round(amount, 2)
            for tier, amount in estimate.total_by_tier.items()
        },
        "count_by_tier": {
            tier.value: count
            for tier, count in estimate.count_by_tier.items()
        },
        "by_division": estimate.by_division,
        "by_cost_category": estimate.by_cost_category,
        "line_items": [li.model_dump() for li in estimate.line_items],
        "rooms":   [r.model_dump() for r in project.rooms],
        "doors":   [d.model_dump() for d in project.doors],
        "windows": [w.model_dump() for w in project.windows],
        "structural": [s.model_dump() for s in project.structural],
        "mep":        [m.model_dump() for m in project.mep],
        "spec_sections": [s.model_dump() for s in project.spec_sections],
        "site": project.site.model_dump(),
        "bid_packages": [p.model_dump() for p in project.bid_packages],
        "supporting_documents": [
            p.model_dump() for p in project.bid_packages
            if p.document_kind == "supporting_document"
        ],
        "scope_matrix": {
            "by_division": {
                div: [p.pdf_name for p in pkgs]
                for div, pkgs in project.scope_matrix.by_division.items()
            },
            "all_alternates": [a.model_dump() for a in project.scope_matrix.all_alternates],
            "coverage_warnings": project.scope_matrix.coverage_warnings,
        },
        "aggregated_inclusions": [
            it.model_dump() for it in getattr(project, "aggregated_inclusions", []) or []
        ],
        "aggregated_exclusions": [
            it.model_dump() for it in getattr(project, "aggregated_exclusions", []) or []
        ],
        "sheet_summaries": project.sheet_summaries,
        "warnings": project.warnings,
        # Phase T9.0 — bid alternates / VE. ``alternates`` holds the
        # priced :class:`AlternateLineEstimate` rows attached to the
        # estimate; the top-level totals mirror the Excel "Bid
        # Alternates" footer rows so JSON consumers (downstream PDF
        # builders deferred to T9.2, Streamlit T9.1 tab) can render
        # the headline numbers without re-walking the list.
        "alternates": [
            a.model_dump() for a in getattr(estimate, "alternates", None) or []
        ],
        "alternates_selected_default": sorted(
            list(getattr(estimate, "alternates_selected_default", None) or set())
        ),
        "alternates_total_additive": estimate.alternates_total_additive,
        "alternates_total_deductive": estimate.alternates_total_deductive,
        "alternates_total_substitution": estimate.alternates_total_substitution,
        "alternates_count_missing": estimate.alternates_count_missing,
        "project_alternates": [
            a.model_dump() for a in getattr(project, "alternates", None) or []
        ],
    }
    return json.dumps(payload, indent=2, default=str)


def save_to_disk(content: bytes | str, path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(content, str):
        path.write_text(content, encoding="utf-8")
    else:
        path.write_bytes(content)
    return path
