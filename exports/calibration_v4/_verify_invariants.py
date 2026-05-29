"""Verify subtotal/grand_total invariants + confidence-band math."""
import json
from pathlib import Path

ROOT = Path(__file__).parent
d = json.load((ROOT / "estimate.json").open("r", encoding="utf-8"))

items = d.get("line_items", [])
priced = [it for it in items if not it.get("suppressed")]
sup = [it for it in items if it.get("suppressed")]

print("=== INVARIANT CHECKS ===")
print(f"  Total line items: {len(items)}")
print(f"  Priced (unsuppressed): {len(priced)}")
print(f"  Suppressed: {len(sup)}")

sum_priced = sum(float(it.get("total_cost") or 0.0) for it in priced)
sum_sup = sum(float(it.get("total_cost") or 0.0) for it in sup)
print(f"\n  Sum of priced.total_cost: ${sum_priced:,.2f}")
print(f"  Sum of suppressed.total_cost: ${sum_sup:,.2f}")
print(f"  reported subtotal:        ${d.get('subtotal'):,.2f}")

# region multiplier
reg = float(d.get("region_multiplier") or 1.0)
print(f"\n  region_multiplier: {reg}")
print(f"  sum_priced (= subtotal): match={abs(sum_priced - d.get('subtotal')) < 1.0}")

# grand_total invariant
st = float(d.get("subtotal"))
c_pct = float(d.get("contingency_pct"))
o_pct = float(d.get("overhead_pct"))
p_pct = float(d.get("profit_pct"))
expected_gt = st * (1 + c_pct / 100) * (1 + o_pct / 100) * (1 + p_pct / 100)
print(f"\n  contingency_pct: {c_pct}")
print(f"  overhead_pct:    {o_pct}")
print(f"  profit_pct:      {p_pct}")
print(f"  expected grand_total: ${expected_gt:,.2f}")
print(f"  reported grand_total: ${d.get('grand_total'):,.2f}")
print(f"  grand_total invariant: match={abs(expected_gt - d.get('grand_total')) < 1.0}")

# Suppressed contribution is $0
print(f"\n  Suppressed lines contribute to subtotal? {'NO' if sum_sup == 0 else 'YES (REGRESSION)'}")

# combined confidence = qty_conf * price_conf invariant
print(f"\n=== COMBINED CONFIDENCE INVARIANT ===")
mismatches = 0
for it in items:
    qc = it.get("confidence")
    pc = it.get("price_confidence")
    # combined isn't directly stored; just verify both exist
    if qc is None or pc is None:
        mismatches += 1
print(f"  Lines missing confidence values: {mismatches}/{len(items)}")

# Confidence band distribution
from collections import Counter
bands = Counter(it.get("cost_band") for it in items)
tiers = Counter(it.get("cost_source_tier") for it in items)
print(f"\n  Bands: {dict(bands)}")
print(f"  Tiers: {dict(tiers)}")

# by_division total
by_div = d.get("by_division", {})
by_div_sum = sum(by_div.values()) if isinstance(by_div, dict) else 0
print(f"\n  by_division sum: ${by_div_sum:,.2f} (must equal subtotal)")
print(f"  by_division == subtotal: {abs(by_div_sum - st) < 1.0}")

# Excel sheets check
print(f"\n=== EXCEL SHEETS ===")
try:
    from openpyxl import load_workbook
    wb = load_workbook(ROOT / "estimate.xlsx", read_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    # Check for Bid Packages + Scope Matrix sheets (B3-3 regression)
    expected_sheets = ["Bid Packages", "Scope Matrix"]
    for name in expected_sheets:
        present = name in wb.sheetnames
        print(f"  '{name}' present: {present}")
except Exception as e:
    print(f"  Excel load failed: {e}")
